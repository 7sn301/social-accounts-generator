"""
🦅 Baseer v2.1.0 - Professional Edition
═══════════════════════════════════════════════════════════════
الميزات (لا تُعرض للمستخدم):
  1. واجهة نظيفة - لا تفاصيل تقنية في Sidebar/Footer
  2. تفاصيل تقنية مرئية بشكل جميل ومرتب
  3. بحث جماعي حتى 50 حساب
  4. استيراد/تصدير Excel
  5. تصميم Responsive للجوال
  6. التحليل الذكي للمنطقة (5 طبقات)
  7. كشف المغتربين
  8. قاعدة 1112 منطقة + 84 عاصمة + 61 مشهور
═══════════════════════════════════════════════════════════════
"""
import streamlit as st
import requests
import re
import json
import time
import random
import io
from datetime import datetime
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).parent / 'data'))
from regions_database import REGIONS_DATABASE, lookup_region
# ❌ v2.1.2 - تعطيل استيراد lookup_capital (ليس مستخدماً بعد حذف تخمين العاصمة)
# from capitals_database import lookup_capital  # DEPRECATED

VERSION = "v2.1.0"  # لا يُعرض في الواجهة

# ═══════════════════════════════════════════════════════════════
# 🎨 إعدادات الصفحة
# ═══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="بَصِير | مولّد معلومات TikTok",
    page_icon="🦅",
    layout="wide",
    initial_sidebar_state="collapsed",  # Sidebar مطوي افتراضياً
)

# ═══════════════════════════════════════════════════════════════
# 🌐 بروكسيات
# ═══════════════════════════════════════════════════════════════
# ✅ v2.1.7-Hardened: Smart Anti-Detection — 22 متصفّح حديث متنوّع
USER_AGENTS = [
    # Chrome على Windows
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
    # Chrome على macOS
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
    # Safari على macOS
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.6 Safari/605.1.15',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15',
    # Firefox على Windows
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:131.0) Gecko/20100101 Firefox/131.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:130.0) Gecko/20100101 Firefox/130.0',
    # Firefox على macOS
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 14.5; rv:131.0) Gecko/20100101 Firefox/131.0',
    # Edge على Windows
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 Edg/130.0.0.0',
    # iPhone Safari
    'Mozilla/5.0 (iPhone; CPU iPhone OS 17_6_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.6 Mobile/15E148 Safari/604.1',
    'Mozilla/5.0 (iPhone; CPU iPhone OS 17_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Mobile/15E148 Safari/604.1',
    'Mozilla/5.0 (iPhone; CPU iPhone OS 16_7_8 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1',
    # iPad Safari
    'Mozilla/5.0 (iPad; CPU OS 17_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Mobile/15E148 Safari/604.1',
    # Android Chrome
    'Mozilla/5.0 (Linux; Android 14; Pixel 8 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Mobile Safari/537.36',
    'Mozilla/5.0 (Linux; Android 13; SM-S918B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Mobile Safari/537.36',
    'Mozilla/5.0 (Linux; Android 14; SM-A546B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Mobile Safari/537.36',
    # Linux Chrome
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:131.0) Gecko/20100101 Firefox/131.0',
    # Opera
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 OPR/116.0.0.0',
]

# ✅ v2.1.7-Light: tikwm.com فقط (النسخة الرسمية)
TIKWM_BASE = 'https://www.tikwm.com'

# ✅ تحسين v2.1.1 - سلسلة بروكسي احتياطية لتجنب فشل jina المفرد
PROXY_CHAIN = [
    {'name': 'jina', 'url': 'https://r.jina.ai/', 'timeout': 15},
    {'name': 'allorigins', 'url': 'https://api.allorigins.win/raw?url=', 'timeout': 18},
    {'name': 'corsproxy', 'url': 'https://corsproxy.io/?', 'timeout': 15},
    {'name': 'codetabs', 'url': 'https://api.codetabs.com/v1/proxy?quest=', 'timeout': 20},
]

# ═══════════════════════════════════════════════════════════════
# 🎯 Patterns
# ═══════════════════════════════════════════════════════════════
PATTERNS = {
    'country_flag': r'(?:🇺🇸|🇬🇧|🇰🇷|🇯🇵|🇨🇳|🇮🇳|🇧🇷|🇿🇦|🇦🇺|🇪🇬|🇸🇦|🇦🇪|🇰🇼|🇶🇦|🇧🇭|🇴🇲|🇯🇴|🇱🇧|🇮🇶|🇾🇪|🇵🇸|🇲🇦|🇩🇿|🇹🇳|🇱🇾|🇸🇩|🇸🇴|🇹🇷|🇮🇷|🇮🇹|🇫🇷|🇩🇪|🇪🇸|🇳🇱|🇷🇺|🇨🇦|🇲🇽|🇦🇷|🇨🇴|🇨🇱|🇵🇪|🇻🇪|🇳🇬|🇰🇪|🇪🇹|🇬🇭|🇹🇭|🇻🇳|🇮🇩|🇲🇾|🇵🇭|🇸🇬|🇵🇰|🇧🇩|🇱🇰|🇳🇿|🇵🇹|🇵🇷|🇹🇿|🇹🇼|🇭🇰)([A-Za-z][A-Za-z\s&\.\(\)\']+?)(?:\n|🌐)',
    'country_globe': r'🌍\s*([A-Za-z][A-Za-z\s&\.\(\)\']+?)(?:\n|🌐|<|$)',
    'language': r'🌐([a-zA-Z]{2,7})',
    'followers': r'([\d,]+)\s*👥\s*Followers|👥 Followers\s*\n*([\d,]+)',
    'following': r'([\d,]+)\s*➕\s*Following|➕ Following\s*\n*([\d,]+)',
    'hearts': r'([\d,]+)\s*❤️\s*Hearts|❤️ Hearts\s*\n*([\d,]+)',
    'videos': r'([\d,]+)\s*🎬\s*Videos|🎬 Videos\s*\n*([\d,]+)',
    'friends': r'([\d,]+)\s*👫\s*Friends|👫 Friends\s*\n*([\d,]+)',
    'user_id': r'User ID:[*\s`]*(\d+)',
    'sec_uid': r'SecUID:[*\s`]*([A-Za-z0-9_-]+)',
    'created': r'Account Created:[*\s]*([^\n]+)',
    'nickname': r'##\s*([^\n#]+)\n\n@',
    'avatar': r'!\[Image[^\]]*\]\((https://[^)]+)\)',
}

BIO_PATTERNS = [
    r'\[📥 Download Videos\][^\n]*\n+([^\n\[]{2,300}?)(?:\n+\[|\n+###|$)',
    r'\*\*📝 About:\*\*\s*([^\n*]+)',
    r'About:\*\*\s*([^\n*]+)',
]

# ═══════════════════════════════════════════════════════════════
# 🌟 قاعدة المشاهير
# ═══════════════════════════════════════════════════════════════
CELEBRITIES = {
    "khaby.lame": {"country": "Italy", "flag": "🇮🇹", "name": "Khabane Lame", "region": "ميلانو"},
    "aboflah": {"country": "Kuwait", "flag": "🇰🇼", "name": "AboFlah", "region": "مدينة الكويت"},
    "shougalhady": {"country": "Kuwait", "flag": "🇰🇼", "name": "Shoug Alhady", "region": "مدينة الكويت"},
    "hayaalshuaibi": {"country": "Kuwait", "flag": "🇰🇼", "name": "Haya Alshuaibi", "region": "مدينة الكويت"},
    "amrdiab": {"country": "Egypt", "flag": "🇪🇬", "name": "Amr Diab", "region": "القاهرة"},
    "mohamedramadanws": {"country": "Egypt", "flag": "🇪🇬", "name": "Mohamed Ramadan", "region": "القاهرة"},
    "aljazeera": {"country": "Qatar", "flag": "🇶🇦", "name": "Al Jazeera", "region": "الدوحة"},
    "blackpinkofficial": {"country": "South Korea", "flag": "🇰🇷", "name": "BLACKPINK", "region": "سيول"},
    "bts_official_bighit": {"country": "South Korea", "flag": "🇰🇷", "name": "BTS", "region": "سيول"},
    "newjeans_official": {"country": "South Korea", "flag": "🇰🇷", "name": "NewJeans", "region": "سيول"},
    "twice_tiktok_official": {"country": "South Korea", "flag": "🇰🇷", "name": "TWICE", "region": "سيول"},
    "ive_official": {"country": "South Korea", "flag": "🇰🇷", "name": "IVE", "region": "سيول"},
    "lesserafim_official": {"country": "South Korea", "flag": "🇰🇷", "name": "LE SSERAFIM", "region": "سيول"},
    "straykids": {"country": "South Korea", "flag": "🇰🇷", "name": "Stray Kids", "region": "سيول"},
    "itzofficial": {"country": "South Korea", "flag": "🇰🇷", "name": "ITZY", "region": "سيول"},
    "bellapoarch": {"country": "Philippines", "flag": "🇵🇭", "name": "Bella Poarch", "region": "مانيلا"},
    "lalisa_manobal": {"country": "Thailand", "flag": "🇹🇭", "name": "Lisa", "region": "بانكوك"},
    "raffinagita1717": {"country": "Indonesia", "flag": "🇮🇩", "name": "Raffi Nagita", "region": "جاكرتا"},
    "priyankachopra": {"country": "India", "flag": "🇮🇳", "name": "Priyanka Chopra", "region": "مومباي"},
    "viratkohli": {"country": "India", "flag": "🇮🇳", "name": "Virat Kohli", "region": "مومباي"},
    "deepikapadukone": {"country": "India", "flag": "🇮🇳", "name": "Deepika Padukone", "region": "مومباي"},
    "alia.bhatt": {"country": "India", "flag": "🇮🇳", "name": "Alia Bhatt", "region": "مومباي"},
    "atifaslam": {"country": "Pakistan", "flag": "🇵🇰", "name": "Atif Aslam", "region": "كراتشي"},
    "son.tung.mtp": {"country": "Vietnam", "flag": "🇻🇳", "name": "Son Tung", "region": "هو تشي مينه"},
    "cristiano": {"country": "Portugal", "flag": "🇵🇹", "name": "Cristiano Ronaldo", "region": "لشبونة"},
    "youness_zarou": {"country": "Germany", "flag": "🇩🇪", "name": "Younes Zarou", "region": "فرانكفورت"},
    "twincoach": {"country": "Germany", "flag": "🇩🇪", "name": "TwinCoach", "region": "برلين"},
    "younesnaffaa": {"country": "Morocco", "flag": "🇲🇦", "name": "Younes Naffa", "region": "الدار البيضاء"},
    "mrbeast": {"country": "United States", "flag": "🇺🇸", "name": "MrBeast", "region": "نيويورك"},
    "charlidamelio": {"country": "United States", "flag": "🇺🇸", "name": "Charli D'Amelio", "region": "لوس أنجلوس"},
    "kingjames": {"country": "United States", "flag": "🇺🇸", "name": "LeBron James", "region": "لوس أنجلوس"},
    "therock": {"country": "United States", "flag": "🇺🇸", "name": "Dwayne Johnson", "region": "لوس أنجلوس"},
    "selenagomez": {"country": "United States", "flag": "🇺🇸", "name": "Selena Gomez", "region": "لوس أنجلوس"},
    "taylorswift": {"country": "United States", "flag": "🇺🇸", "name": "Taylor Swift", "region": "ناشفيل"},
    "kimkardashian": {"country": "United States", "flag": "🇺🇸", "name": "Kim Kardashian", "region": "لوس أنجلوس"},
    "kyliejenner": {"country": "United States", "flag": "🇺🇸", "name": "Kylie Jenner", "region": "لوس أنجلوس"},
    "billieeilish": {"country": "United States", "flag": "🇺🇸", "name": "Billie Eilish", "region": "لوس أنجلوس"},
    "arianagrande": {"country": "United States", "flag": "🇺🇸", "name": "Ariana Grande", "region": "لوس أنجلوس"},
    "zendaya": {"country": "United States", "flag": "🇺🇸", "name": "Zendaya", "region": "لوس أنجلوس"},
    "addisonre": {"country": "United States", "flag": "🇺🇸", "name": "Addison Rae", "region": "لوس أنجلوس"},
    "willsmith": {"country": "United States", "flag": "🇺🇸", "name": "Will Smith", "region": "لوس أنجلوس"},
    "drake": {"country": "Canada", "flag": "🇨🇦", "name": "Drake", "region": "تورنتو"},
    "shawnmendes": {"country": "Canada", "flag": "🇨🇦", "name": "Shawn Mendes", "region": "تورنتو"},
    "justinbieber": {"country": "Canada", "flag": "🇨🇦", "name": "Justin Bieber", "region": "تورنتو"},
    "shakira": {"country": "Colombia", "flag": "🇨🇴", "name": "Shakira", "region": "بارانكويلا"},
    "jbalvin": {"country": "Colombia", "flag": "🇨🇴", "name": "J Balvin", "region": "ميديلين"},
    "karolg": {"country": "Colombia", "flag": "🇨🇴", "name": "Karol G", "region": "ميديلين"},
    "messi": {"country": "Argentina", "flag": "🇦🇷", "name": "Lionel Messi", "region": "بوينس آيرس"},
    "tini": {"country": "Argentina", "flag": "🇦🇷", "name": "TINI", "region": "بوينس آيرس"},
    "anitta": {"country": "Brazil", "flag": "🇧🇷", "name": "Anitta", "region": "ريو دي جانيرو"},
    "neymarjr": {"country": "Brazil", "flag": "🇧🇷", "name": "Neymar Jr", "region": "ساو باولو"},
    "lelepons": {"country": "Venezuela", "flag": "🇻🇪", "name": "Lele Pons", "region": "كاراكاس"},
    "badbunny": {"country": "Puerto Rico", "flag": "🇵🇷", "name": "Bad Bunny", "region": "بورتو ريكو"},
    "wizkidayo": {"country": "Nigeria", "flag": "🇳🇬", "name": "Wizkid", "region": "لاغوس"},
    "burnaboygram": {"country": "Nigeria", "flag": "🇳🇬", "name": "Burna Boy", "region": "لاغوس"},
    "davidoofficial": {"country": "Nigeria", "flag": "🇳🇬", "name": "Davido", "region": "لاغوس"},
    "tiwasavage": {"country": "Nigeria", "flag": "🇳🇬", "name": "Tiwa Savage", "region": "لاغوس"},
    "trevornoah": {"country": "South Africa", "flag": "🇿🇦", "name": "Trevor Noah", "region": "جوهانسبرغ"},
    "blackcoffeeofficial": {"country": "South Africa", "flag": "🇿🇦", "name": "Black Coffee", "region": "جوهانسبرغ"},
    "diamondplatnumz": {"country": "Tanzania", "flag": "🇹🇿", "name": "Diamond Platnumz", "region": "دار السلام"},
    "chrishemsworth": {"country": "Australia", "flag": "🇦🇺", "name": "Chris Hemsworth", "region": "سيدني"},
}

LOCAL_VERIFIED_DB = {
    "zahranabill1": {"country": "Kuwait", "flag": "🇰🇼", "name": "Zahra Nabill", "region": "مدينة الكويت"},
}

# ═══════════════════════════════════════════════════════════════
# 🗣️ خرائط
# ═══════════════════════════════════════════════════════════════
# ✅ v2.1.5 - عرض اللغة بصيغة عربية+إنجليزية واضحة
LANGUAGE_DISPLAY_AR = {
    'ar': 'العربية', 'en': 'الإنجليزية', 'fr': 'الفرنسية', 'es': 'الإسبانية',
    'de': 'الألمانية', 'it': 'الإيطالية', 'pt': 'البرتغالية', 'ru': 'الروسية',
    'tr': 'التركية', 'ko': 'الكورية', 'ja': 'اليابانية', 'zh': 'الصينية',
    'th': 'التايلاندية', 'vi': 'الفيتنامية', 'id': 'الإندونيسية',
    'hi': 'الهندية', 'ur': 'الأردية', 'fa': 'الفارسية', 'nl': 'الهولندية',
    'sv': 'السويدية', 'no': 'النرويجية', 'fi': 'الفنلندية', 'da': 'الدنماركية',
    'pl': 'البولندية', 'cs': 'التشيكية', 'el': 'اليونانية', 'he': 'العبرية',
    'ms': 'الماليزية', 'tl': 'الفلبينية', 'bn': 'البنغالية', 'ta': 'التاميلية',
}

def _format_language_label(lang_code):
    """✅ v2.1.5 - تنسيق اللغة بصيغة 'en — الإنجليزية' لإخفاء الرموز التقنية"""
    if not lang_code:
        return '—'
    code = str(lang_code).strip().lower()[:2]
    if not code:
        return '—'
    ar_name = LANGUAGE_DISPLAY_AR.get(code)
    if ar_name:
        return f"{code} — {ar_name}"
    return code

# ✅ إصلاح v2.1.1 - اللغة العربية لا تحدد دولة واحدة
# (تغطي 22+ دولة) - فقط تستخدم للتحقق من ترابط الجنسية
LANGUAGE_TO_COUNTRY = {
    'ar': (None, ['Saudi Arabia', 'United Arab Emirates', 'Egypt', 'Kuwait', 'Qatar', 'Bahrain', 'Oman', 'Jordan', 'Lebanon', 'Iraq', 'Yemen', 'Palestine', 'Morocco', 'Algeria', 'Tunisia', 'Libya', 'Sudan']),
    'ko': ('South Korea', ['South Korea', 'Korea']),
    'ja': ('Japan', ['Japan']),
    'zh': ('China', ['China', 'Taiwan', 'Hong Kong', 'Singapore']),
    'th': ('Thailand', ['Thailand']), 'vi': ('Vietnam', ['Vietnam']),
    'id': ('Indonesia', ['Indonesia']), 'tr': ('Turkey', ['Turkey']),
    'hi': ('India', ['India']), 'ur': ('Pakistan', ['Pakistan']),
    'pt': ('Brazil', ['Brazil', 'Portugal']),
    'es': ('Spain', ['Spain', 'Mexico', 'Argentina', 'Colombia', 'Chile', 'Peru', 'Venezuela']),
    'fr': ('France', ['France', 'Belgium', 'Canada', 'Switzerland']),
    'de': ('Germany', ['Germany', 'Austria', 'Switzerland']),
    'it': ('Italy', ['Italy']), 'ru': ('Russia', ['Russia', 'Kazakhstan']),
    'en': ('United States', ['United States', 'United Kingdom', 'Canada', 'Australia', 'New Zealand']),
}

USERNAME_KEYWORDS = {
    'japan': 'Japan', 'tokyo': 'Japan', 'korea': 'South Korea', 'seoul': 'South Korea',
    'kpop': 'South Korea', 'thai': 'Thailand', 'bangkok': 'Thailand',
    'viet': 'Vietnam', 'indo': 'Indonesia', 'jakarta': 'Indonesia',
    'malay': 'Malaysia', 'manila': 'Philippines', 'india': 'India',
    'mumbai': 'India', 'pakistan': 'Pakistan',
    'qatar': 'Qatar', 'kuwait': 'Kuwait', 'saudi': 'Saudi Arabia', 'riyadh': 'Saudi Arabia',
    'emirates': 'United Arab Emirates', 'dubai': 'United Arab Emirates',
    'egypt': 'Egypt', 'cairo': 'Egypt', 'morocco': 'Morocco',
    'paris': 'France', 'berlin': 'Germany', 'madrid': 'Spain',
    'italy': 'Italy', 'brazil': 'Brazil', 'argentina': 'Argentina',
    'mexico': 'Mexico', 'nigeria': 'Nigeria', 'australia': 'Australia',
}
USERNAME_KEYWORDS_SORTED = sorted(USERNAME_KEYWORDS.items(), key=lambda x: -len(x[0]))

FLAG_EMOJI_TO_COUNTRY = {
    '🇸🇦': 'Saudi Arabia', '🇦🇪': 'United Arab Emirates', '🇪🇬': 'Egypt',
    '🇰🇼': 'Kuwait', '🇶🇦': 'Qatar', '🇧🇭': 'Bahrain', '🇴🇲': 'Oman',
    '🇯🇴': 'Jordan', '🇱🇧': 'Lebanon', '🇮🇶': 'Iraq', '🇾🇪': 'Yemen',
    '🇵🇸': 'Palestine', '🇲🇦': 'Morocco', '🇩🇿': 'Algeria', '🇹🇳': 'Tunisia',
    '🇱🇾': 'Libya', '🇸🇩': 'Sudan',
    '🇰🇷': 'South Korea', '🇯🇵': 'Japan', '🇨🇳': 'China', '🇹🇼': 'Taiwan',
    '🇭🇰': 'Hong Kong', '🇸🇬': 'Singapore', '🇹🇭': 'Thailand', '🇻🇳': 'Vietnam',
    '🇮🇩': 'Indonesia', '🇲🇾': 'Malaysia', '🇵🇭': 'Philippines',
    '🇮🇳': 'India', '🇵🇰': 'Pakistan', '🇧🇩': 'Bangladesh', '🇱🇰': 'Sri Lanka',
    '🇺🇸': 'United States', '🇬🇧': 'United Kingdom', '🇨🇦': 'Canada',
    '🇫🇷': 'France', '🇩🇪': 'Germany', '🇮🇹': 'Italy', '🇪🇸': 'Spain',
    '🇳🇱': 'Netherlands', '🇵🇹': 'Portugal', '🇹🇷': 'Turkey', '🇷🇺': 'Russia',
    '🇲🇽': 'Mexico', '🇧🇷': 'Brazil', '🇦🇷': 'Argentina', '🇨🇴': 'Colombia',
    '🇨🇱': 'Chile', '🇵🇪': 'Peru', '🇻🇪': 'Venezuela', '🇵🇷': 'Puerto Rico',
    '🇳🇬': 'Nigeria', '🇿🇦': 'South Africa', '🇰🇪': 'Kenya', '🇬🇭': 'Ghana',
    '🇹🇿': 'Tanzania', '🇪🇹': 'Ethiopia', '🇦🇺': 'Australia', '🇳🇿': 'New Zealand',
}

# ✅ تطوير #4 v2.1.1 - إضافة كشف العربية (بدون تخصيص دولة واحدة)
BIO_SCRIPT_TO_COUNTRY = [
    (r'[\uAC00-\uD7AF]', 'South Korea', 'Hangul'),
    (r'[\u3040-\u309F]', 'Japan', 'Hiragana'),
    (r'[\u30A0-\u30FF]', 'Japan', 'Katakana'),
    (r'[\u4E00-\u9FFF]', 'China', 'Chinese'),
    (r'[\u0E00-\u0E7F]', 'Thailand', 'Thai'),
    (r'[\u0900-\u097F]', 'India', 'Devanagari'),
    # العربية لا تحدد دولة واحدة (22+ دولة)
    # تستخدم فقط للإشارة أن المحتوى عربي - تترك التصحيح لبيانات TikMatrix
    (r'[\u0600-\u06FF]', None, 'Arabic'),
]

TLD_TO_COUNTRY = {
    '.sa': 'Saudi Arabia', '.ae': 'United Arab Emirates', '.kw': 'Kuwait',
    '.qa': 'Qatar', '.eg': 'Egypt', '.ma': 'Morocco', '.dz': 'Algeria',
    '.kr': 'South Korea', '.jp': 'Japan', '.fr': 'France', '.de': 'Germany',
    '.it': 'Italy', '.es': 'Spain', '.br': 'Brazil', '.au': 'Australia',
}

AMBIGUOUS_TLDS = {'.tn', '.tv', '.io', '.ai', '.co', '.me', '.ly', '.fm'}
EXPANDED_SUSPICIOUS = {'Turks and Caicos Islands', 'Norway', 'Sweden', 'Finland', 'Puerto Rico', 'Sri Lanka', 'Netherlands', 'Ireland', 'Iraq'}
TIKTOK_SERVER_COUNTRIES = {'United States', 'United Kingdom'}
GLOBAL_LANGUAGES = {'en'}
TRUSTED_TIKMATRIX_COUNTRIES = set(FLAG_EMOJI_TO_COUNTRY.values())

MISLEADING_PATTERNS = {
    'newyork': 'United States', 'nyc': 'United States', 'losangeles': 'United States',
    'london': 'United Kingdom', 'paris': 'France', 'berlin': 'Germany',
    'roma': 'Italy', 'madrid': 'Spain',
}

# ═══════════════════════════════════════════════════════════════
# 🦅 المحرك
# ═══════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════
# ✅ v2.1.6 طارئ - مصادر بيانات جديدة لاستعادة دقة الموقع
# ═══════════════════════════════════════════════════════════════
# ✅ v2.1.7 - خريطة ISO codes شاملة (250+ دولة وفق ISO 3166-1 alpha-2)
REGION_ISO_TO_COUNTRY = {
    # العربية
    'SA': 'Saudi Arabia', 'AE': 'United Arab Emirates', 'KW': 'Kuwait',
    'QA': 'Qatar', 'BH': 'Bahrain', 'OM': 'Oman', 'YE': 'Yemen',
    'JO': 'Jordan', 'LB': 'Lebanon', 'IQ': 'Iraq', 'PS': 'Palestine',
    'EG': 'Egypt', 'MA': 'Morocco', 'DZ': 'Algeria', 'TN': 'Tunisia',
    'LY': 'Libya', 'SD': 'Sudan', 'SO': 'Somalia', 'MR': 'Mauritania',
    'DJ': 'Djibouti', 'KM': 'Comoros',
    # أوروبا وأمريكا الشمالية وأستراليا
    'US': 'United States', 'GB': 'United Kingdom', 'UK': 'United Kingdom',
    'CA': 'Canada', 'AU': 'Australia', 'NZ': 'New Zealand', 'IE': 'Ireland',
    'FR': 'France', 'DE': 'Germany', 'IT': 'Italy', 'ES': 'Spain',
    'PT': 'Portugal', 'NL': 'Netherlands', 'BE': 'Belgium', 'CH': 'Switzerland',
    'AT': 'Austria', 'SE': 'Sweden', 'NO': 'Norway', 'FI': 'Finland',
    'DK': 'Denmark', 'PL': 'Poland', 'CZ': 'Czech Republic', 'SK': 'Slovakia',
    'HU': 'Hungary', 'RO': 'Romania', 'BG': 'Bulgaria', 'GR': 'Greece',
    'HR': 'Croatia', 'SI': 'Slovenia', 'RS': 'Serbia', 'BA': 'Bosnia and Herzegovina',
    'MK': 'North Macedonia', 'AL': 'Albania', 'ME': 'Montenegro', 'XK': 'Kosovo',
    'EE': 'Estonia', 'LV': 'Latvia', 'LT': 'Lithuania', 'IS': 'Iceland',
    'LU': 'Luxembourg', 'MT': 'Malta', 'CY': 'Cyprus', 'MC': 'Monaco',
    'AD': 'Andorra', 'SM': 'San Marino', 'VA': 'Vatican City', 'LI': 'Liechtenstein',
    # روسيا ودول السوفيت سابقاً
    'RU': 'Russia', 'UA': 'Ukraine', 'BY': 'Belarus', 'MD': 'Moldova',
    'GE': 'Georgia', 'AM': 'Armenia', 'AZ': 'Azerbaijan', 'KZ': 'Kazakhstan',
    'UZ': 'Uzbekistan', 'KG': 'Kyrgyzstan', 'TJ': 'Tajikistan', 'TM': 'Turkmenistan',
    # تركيا وإسرائيل وإيران
    'TR': 'Turkey', 'IL': 'Israel', 'IR': 'Iran',
    # جنوب آسيا
    'IN': 'India', 'PK': 'Pakistan', 'BD': 'Bangladesh', 'LK': 'Sri Lanka',
    'NP': 'Nepal', 'AF': 'Afghanistan', 'BT': 'Bhutan', 'MV': 'Maldives',
    # شرق آسيا
    'CN': 'China', 'JP': 'Japan', 'KR': 'South Korea', 'KP': 'North Korea',
    'TW': 'Taiwan', 'HK': 'Hong Kong', 'MO': 'Macau', 'MN': 'Mongolia',
    # جنوب شرق آسيا
    'SG': 'Singapore', 'MY': 'Malaysia', 'ID': 'Indonesia', 'TH': 'Thailand',
    'VN': 'Vietnam', 'PH': 'Philippines', 'MM': 'Myanmar', 'KH': 'Cambodia',
    'LA': 'Laos', 'BN': 'Brunei', 'TL': 'Timor-Leste',
    # أمريكا اللاتينية
    'BR': 'Brazil', 'MX': 'Mexico', 'AR': 'Argentina', 'CL': 'Chile',
    'CO': 'Colombia', 'PE': 'Peru', 'VE': 'Venezuela', 'UY': 'Uruguay',
    'PY': 'Paraguay', 'BO': 'Bolivia', 'EC': 'Ecuador', 'GY': 'Guyana',
    'SR': 'Suriname', 'GT': 'Guatemala', 'HN': 'Honduras', 'SV': 'El Salvador',
    'NI': 'Nicaragua', 'CR': 'Costa Rica', 'PA': 'Panama', 'DO': 'Dominican Republic',
    'CU': 'Cuba', 'HT': 'Haiti', 'JM': 'Jamaica', 'PR': 'Puerto Rico',
    'TT': 'Trinidad and Tobago', 'BS': 'Bahamas', 'BB': 'Barbados', 'BZ': 'Belize',
    # أفريقيا
    'NG': 'Nigeria', 'KE': 'Kenya', 'ET': 'Ethiopia', 'GH': 'Ghana',
    'ZA': 'South Africa', 'TZ': 'Tanzania', 'UG': 'Uganda', 'CI': 'Ivory Coast',
    'SN': 'Senegal', 'CM': 'Cameroon', 'ML': 'Mali', 'BF': 'Burkina Faso',
    'NE': 'Niger', 'TD': 'Chad', 'AO': 'Angola', 'MZ': 'Mozambique',
    'ZW': 'Zimbabwe', 'ZM': 'Zambia', 'MW': 'Malawi', 'BW': 'Botswana',
    'NA': 'Namibia', 'LS': 'Lesotho', 'SZ': 'Eswatini', 'MG': 'Madagascar',
    'MU': 'Mauritius', 'SC': 'Seychelles', 'RW': 'Rwanda', 'BI': 'Burundi',
    'CD': 'DR Congo', 'CG': 'Congo', 'GA': 'Gabon', 'GQ': 'Equatorial Guinea',
    'CF': 'Central African Republic', 'SS': 'South Sudan', 'ER': 'Eritrea',
    'LR': 'Liberia', 'SL': 'Sierra Leone', 'GN': 'Guinea', 'GM': 'Gambia',
    'GW': 'Guinea-Bissau', 'CV': 'Cape Verde', 'TG': 'Togo', 'BJ': 'Benin',
    'ST': 'Sao Tome and Principe',
    # أوقيانوسيا
    'FJ': 'Fiji', 'PG': 'Papua New Guinea', 'SB': 'Solomon Islands',
    'VU': 'Vanuatu', 'WS': 'Samoa', 'TO': 'Tonga', 'KI': 'Kiribati',
    'TV': 'Tuvalu', 'NR': 'Nauru', 'PW': 'Palau', 'FM': 'Micronesia',
    'MH': 'Marshall Islands',
}

# ✅ v2.1.7 - علم تلقائي من ISO code (تحويل بحرفين إلى emoji regional indicator)
def _iso_to_flag(iso_code):
    """تحويل ISO-3166 alpha-2 إلى emoji علم تلقائياً - يغطي 250+ دولة"""
    if not iso_code or len(iso_code) != 2:
        return ''
    try:
        c1, c2 = iso_code.upper()
        return chr(0x1F1E6 + (ord(c1) - ord('A'))) + chr(0x1F1E6 + (ord(c2) - ord('A')))
    except Exception:
        return ''

# خريطة توليد علم من اسم الدولة — تبنى تلقائياً من REGION_ISO_TO_COUNTRY
COUNTRY_TO_FLAG_EMOJI = {country: _iso_to_flag(iso) for iso, country in REGION_ISO_TO_COUNTRY.items()}
# تجاوزات UK (لأن لدينا GB وUK لنفس الدولة)
COUNTRY_TO_FLAG_EMOJI['United Kingdom'] = '🇬🇧'

# ═════════════════════════════════════════════════════════════
# ✅ v2.1.7-Light-Fix2 — Optimized Throttling + Adaptive Burst
# ═════════════════════════════════════════════════════════════
_TIKWM_LAST_CALL = {'t': 0.0}
_RATELIMIT_BURST = {'until': 0.0}  # ✨ متى ينتهي وضع الحماية بعد 429

def _is_in_burst_mode():
    """يفحص إذا كنا في فترة حماية بعد 429 سابق"""
    return time.time() < _RATELIMIT_BURST['until']

def _activate_burst_protection():
    """يُفعَّل لمدة 60 ث بعد أي 429"""
    _RATELIMIT_BURST['until'] = time.time() + 60.0

def _tikwm_rate_limit_delay():
    """✅ v2.1.7-Light-Fix2 - Adaptive jitter:
    - الوضع العادي: 0.8-1.3ث
    - وضع الحماية (بعد 429): 1.6-2.6ث لمدة 60ث"""
    now = time.time()
    elapsed = now - _TIKWM_LAST_CALL['t']
    if _is_in_burst_mode():
        min_delay = random.uniform(1.6, 2.6)
    else:
        min_delay = random.uniform(0.8, 1.3)
    if elapsed < min_delay:
        time.sleep(min_delay - elapsed)
    _TIKWM_LAST_CALL['t'] = time.time()

# ═════════════════════════════════════════════════════════════
# ✅ v2.1.7-Light — Retry بسيط فقط (بدون headers ثقيلة)
# ═════════════════════════════════════════════════════════════

def fetch_user_tikwm(username):
    """✅ v2.1.7-Light - tikwm مع Retry بسيط (بدون Origin/Referer)"""
    url = f"{TIKWM_BASE}/api/user/info?unique_id={username}"
    start = time.time()
    last_err = None
    for attempt in range(3):
        try:
            _tikwm_rate_limit_delay()
            # ✅ headers بسيطة — فقط User-Agent + Accept (لا Origin/Referer)
            headers = {
                'User-Agent': random.choice(USER_AGENTS),
                'Accept': 'application/json',
            }
            r = requests.get(url, headers=headers, timeout=12 + attempt * 3)
            if r.status_code == 200:
                try:
                    j = r.json()
                    elapsed = time.time() - start
                    # ✅ v2.1.7-Light-Fix1 — توحيد المفاتيح (json/proxy/time)
                    if j.get('code') == 0 and j.get('data', {}).get('user', {}).get('uniqueId'):
                        return {'success': True, 'json': j,
                                'proxy': f'tikwm_attempt_{attempt+1}',
                                'time': round(elapsed, 2)}
                    # لا user.uniqueId → إعادة المحاولة
                    last_err = 'invalid_payload'
                    continue
                except Exception as e:
                    last_err = f'json_parse_failed: {e}'
                    continue
            if r.status_code == 429:
                # ✅ v2.1.7-Light-Fix2 - تفعيل burst protection + backoff تدريجي
                _activate_burst_protection()
                time.sleep(1.5 * (attempt + 1) + random.uniform(0.3, 0.8))
                last_err = 'rate_limit_429'
                continue
            last_err = f'http_{r.status_code}'
        except requests.exceptions.Timeout:
            last_err = 'timeout'
            continue
        except Exception as e:
            last_err = str(e)[:80]
            continue
    # التراجع الأخير للطريقة الكلاسيكية (للتوافق)
    try:
        url = f"https://www.tikwm.com/api/user/info?unique_id={username}"
        headers = {'User-Agent': random.choice(USER_AGENTS), 'Accept': 'application/json'}
        r = requests.get(url, headers=headers, timeout=12)
        elapsed = time.time() - start
        if r.status_code == 200:
            j = r.json()
            if j.get('code') == 0 and j.get('data', {}).get('user', {}).get('uniqueId'):
                return {'success': True, 'json': j, 'proxy': 'tikwm', 'time': round(elapsed, 2)}
    except Exception:
        pass
    # ✅ v2.1.7-Light-Fix1 — إرجاع فشل موحد بكل المفاتيح
    return {'success': False, 'json': None, 'proxy': None,
            'time': 0, 'error': last_err or 'unknown'}

def fetch_user_region_tikwm(username):
    """✅ v2.1.7-Light-Fix3 - جلب region + createTime لآخر 5 فيديوهات
    يُرجع dict {success, region_iso, actual_residence, previous_residence,
                  confidence, residence_type, timezone_match, videos_count, source}
    منطق الإقامة الفعلية = region + تتابع زمني + timezone (فقط)
    احتراماً للقيود — لا مدن، لا هاشتاجات، لا لهجة، لا تعليقات، لا موسيقى"""
    url = f"https://www.tikwm.com/api/user/posts?unique_id={username}&count=5"
    for attempt in range(3):
        try:
            _tikwm_rate_limit_delay()
            headers = {
                'User-Agent': random.choice(USER_AGENTS),
                'Accept': 'application/json',
            }
            r = requests.get(url, headers=headers, timeout=15)
            if r.status_code == 200:
                j = r.json()
                videos = (j.get('data') or {}).get('videos') or []
                if not videos:
                    return {'success': False, 'region_iso': None,
                            'videos_count': 0, 'source': 'tikwm_empty'}

                # ترتيب حسب createTime DESC (الأحدث أولاً)
                videos_sorted = sorted(
                    videos,
                    key=lambda v: v.get('create_time') or 0,
                    reverse=True
                )

                # استخراج regions مع التواريخ
                regions_seq = []
                times_seq = []
                for v in videos_sorted:
                    reg = (v.get('region') or '').strip().upper()
                    ts = v.get('create_time') or 0
                    if reg and len(reg) == 2 and reg.isalpha():
                        regions_seq.append(reg)
                        times_seq.append(ts)

                if not regions_seq:
                    return {'success': False, 'region_iso': None,
                            'videos_count': len(videos), 'source': 'tikwm_no_region'}

                # تحليل الإقامة الفعلية
                analysis = detect_actual_residence(regions_seq, times_seq)

                # ✅ Fix3.1 - حساب توزيع المناطق
                from collections import Counter as _Counter
                region_distribution = dict(_Counter(regions_seq))

                return {
                    'success': True,
                    'region_iso': regions_seq[0],  # آخر فيديو (التوافق مع Fix2)
                    'actual_residence': analysis['actual_residence'],
                    'previous_residence': analysis['previous_residence'],
                    'residence_confidence': analysis['confidence'],
                    'residence_type': analysis['residence_type'],
                    'timezone_match': analysis['timezone_match'],
                    'regions_sequence': regions_seq,
                    'region_distribution': region_distribution,
                    'videos_count': len(regions_seq),
                    'source': 'tikwm_fix3',
                }
            elif r.status_code == 429:
                _activate_burst_protection()
                time.sleep(1.2 + attempt * 1.0)
                continue
        except Exception:
            time.sleep(1.0)
            continue
    return {'success': False, 'region_iso': None, 'videos_count': 0,
            'actual_residence': None, 'residence_confidence': 0}


def _infer_timezone_from_hours(timestamps):
    """✅ v2.1.7-Light-Fix3 - يستنتج timezone من ساعات النشر UTC
    يُرجع ISO للمنطقة الأرجح، أو None
    المنطق — البشر ينشرون في ساعات يقظتهم المحلية (غالباً 09:00-23:00 محلياً)"""
    if not timestamps or len(timestamps) < 3:
        return None

    from datetime import datetime as _dt
    hours_utc = []
    for ts in timestamps:
        try:
            h = _dt.utcfromtimestamp(int(ts)).hour
            hours_utc.append(h)
        except (ValueError, TypeError, OSError):
            continue

    if len(hours_utc) < 3:
        return None

    # متوسط ساعة النشر UTC
    avg_hour = sum(hours_utc) / len(hours_utc)

    # خرائط تقريبية لـ timezone بناءً على ساعات النشاط (15:00-22:00 محلي)
    # إذا أعلى نشاط للساعة X UTC فالتوقيت المحلي = X + offset
    # ساعة الذروة المحلية ~ 19:00، لذا offset = 19 - avg_hour_utc
    offset = round(19 - avg_hour)
    # عادل وتغليف (offset بين -12 إلى +14)
    if offset > 14:
        offset -= 24
    elif offset < -12:
        offset += 24

    # خريطة offset → dominant ISO
    timezone_map = {
        -8: 'US', -7: 'US', -6: 'US', -5: 'US', -4: 'US',
        -3: 'BR', -2: 'BR', -1: 'BR',
        0: 'GB', 1: 'DE', 2: 'EG',
        3: 'SA', 4: 'AE', 5: 'PK', 5.5: 'IN',
        6: 'BD', 7: 'TH', 8: 'CN', 9: 'JP', 10: 'AU',
    }
    return timezone_map.get(offset)


def detect_actual_residence(regions_seq, times_seq):
    """✅ v2.1.7-Light-Fix3 - تحديد الإقامة الفعلية وفق 3 إشارات:
    1. region آخر 5 فيديوهات (70%)
    2. التتابع الزمني — 3 متتالية (20%)
    3. timezone من createTime (10%)

    احتراماً لرفض القائد: لا مدن، لا هاشتاجات، لا لهجة، لا تعليقات، لا موسيقى"""
    from collections import Counter

    if not regions_seq:
        return {'actual_residence': None, 'previous_residence': None,
                'confidence': 0, 'residence_type': 'unknown',
                'timezone_match': False}

    # 1️⃣ فحص آخر 3 فيديوهات
    last_3 = regions_seq[:3]
    last_5 = regions_seq[:5]
    counter_5 = Counter(last_5)
    top_country, top_count = counter_5.most_common(1)[0]

    # 2️⃣ تحديد الإقامة الفعلية + الثقة
    if len(set(last_3)) == 1 and last_3[0]:
        # 3 متتالية متطابقة
        actual = last_3[0]
        confidence = 85
        rtype = 'ثابت' if top_count >= 4 else 'انتقال حديث'
    elif top_count >= 4:
        # 4 من 5 في نفس الدولة
        actual = top_country
        confidence = 80
        rtype = 'ثابت'
    elif top_count >= 3:
        # 3 من 5
        actual = top_country
        confidence = 70
        rtype = 'غالبًا ثابت'
    else:
        # توزيع متفرّق
        actual = regions_seq[0]
        confidence = 50
        rtype = 'مسافر/متنقّل'

    # 3️⃣ كشف الإقامة السابقة (إذا اختلفت)
    previous = None
    if len(regions_seq) >= 5:
        older_regions = regions_seq[3:]  # الفيديوهات الأقدم
        if older_regions:
            older_top = Counter(older_regions).most_common(1)[0][0]
            if older_top and older_top != actual:
                previous = older_top

    # 4️⃣ تأكيد بواسطة timezone
    tz_iso = _infer_timezone_from_hours(times_seq)
    timezone_match = bool(tz_iso and tz_iso == actual)
    if timezone_match:
        confidence = min(confidence + 10, 95)

    return {
        'actual_residence': actual,
        'previous_residence': previous,
        'confidence': confidence,
        'residence_type': rtype,
        'timezone_match': timezone_match,
        'timezone_iso': tz_iso,
        'last_3_distribution': dict(Counter(last_3)),
        'last_5_distribution': dict(counter_5),
    }

@st.cache_data(ttl=300, show_spinner=False)
def fetch_user(username):
    """✅ v2.1.6 - جلب هجين: tikwm أساسي + jina/tikmatrix احتياطي
    يُرجع dict موحد فيه success + content (إذا jina) أو tikwm_json (إذا tikwm) + region_iso"""
    # الطبقة الأساسية: tikwm للبيانات الأساسية
    primary = fetch_user_tikwm(username)
    region_data = {'region_iso': None}
    # ✅ v2.1.7-Light-Fix1 — دفاعي: .get() بدل [] لتجنب KeyError
    if primary.get('success'):
        # جلب region من posts endpoint بشكل متتابع
        region_data = fetch_user_region_tikwm(username)
        return {
            'success': True,
            'tikwm_json': primary.get('json'),
            'region_iso': region_data.get('region_iso'),
            # ✅ Fix3.1 - تمرير جميع حقول الإقامة الفعلية للأعلى
            'actual_residence': region_data.get('actual_residence'),
            'previous_residence': region_data.get('previous_residence'),
            'residence_confidence': region_data.get('residence_confidence', 0),
            'residence_type': region_data.get('residence_type', '—'),
            'timezone_match': region_data.get('timezone_match'),
            'regions_sequence': region_data.get('regions_sequence') or [],
            'region_distribution': region_data.get('region_distribution') or {},
            'videos_count': region_data.get('videos_count', 0),
            'region_source': region_data.get('source'),
            'content': None,
            'proxy': primary.get('proxy', 'tikwm'),
            'time': primary.get('time', 0),
        }
    # التغليف الاحتياطي (jina + tikmatrix) - لحسابات لا تظهر في tikwm
    target = f"https://user.tikmatrix.com/?username={username}"
    for proxy in PROXY_CHAIN:
        url = proxy['url'] + target
        headers = {'User-Agent': random.choice(USER_AGENTS), 'Accept': 'text/html,application/xhtml+xml'}
        if proxy['name'] == 'jina':
            headers['X-Return-Format'] = 'markdown'
        try:
            start = time.time()
            r = requests.get(url, headers=headers, timeout=proxy['timeout'])
            elapsed = time.time() - start
            if r.status_code == 200 and len(r.text) > 1000:
                return {
                    'success': True, 'content': r.text, 'tikwm_json': None,
                    'region_iso': None, 'proxy': proxy['name'], 'time': round(elapsed, 2),
                }
        except Exception:
            continue
    return {'success': False, 'content': None, 'tikwm_json': None, 'region_iso': None, 'proxy': None, 'time': 0}


def _detect_language_from_text(text):
    """✅ v2.1.6 - استنتاج رمز اللغة من تحليل حروف bio/nickname
    يعتمد على أغلبية الأحرف دون تخمين - يرجع None إذا لم توجد أغلبية واضحة"""
    if not text:
        return None
    text = str(text)
    arabic = len(re.findall(r'[\u0600-\u06FF\u0750-\u077F]', text))
    latin = len(re.findall(r'[A-Za-z]', text))
    cjk = len(re.findall(r'[\u4E00-\u9FFF]', text))
    hangul = len(re.findall(r'[\uAC00-\uD7AF]', text))
    hiragana_katakana = len(re.findall(r'[\u3040-\u30FF]', text))
    cyrillic = len(re.findall(r'[\u0400-\u04FF]', text))
    devanagari = len(re.findall(r'[\u0900-\u097F]', text))
    thai = len(re.findall(r'[\u0E00-\u0E7F]', text))
    total = arabic + latin + cjk + hangul + hiragana_katakana + cyrillic + devanagari + thai
    if total < 3:
        return None  # لا تخمين بلا دليل
    scores = [(arabic, 'ar'), (hangul, 'ko'), (hiragana_katakana, 'ja'),
              (cjk, 'zh'), (cyrillic, 'ru'), (devanagari, 'hi'), (thai, 'th'),
              (latin, 'en')]
    scores.sort(reverse=True)
    top, code = scores[0]
    if top / total >= 0.5:
        return code
    return None


# ✅ تحسين v2.1.1 - كشف ديناميكي لحسابات TikMatrix التشغيلية
TIKMATRIX_FALLBACK_ACCOUNTS = {
    'tikmatrix001', 'tikmatrixphonefarm', 'tikmatrix002', 'tikmatrix003',
    'tikmatrix_official', 'tikmatrixbot', 'tikmatrix_demo',
}

TIKMATRIX_FALLBACK_PATTERN = re.compile(r'^tikmatrix[\w_\-]*$', re.IGNORECASE)

def verify_username_match(content, requested):
    m = re.search(r'@([\w\.]+)', content)
    if m:
        actual = m.group(1).lower().strip()
        requested_lower = requested.lower().strip()
        # تحقق 1: قائمة بيضاء لحسابات fallback
        if actual in TIKMATRIX_FALLBACK_ACCOUNTS:
            return False, actual
        # تحقق 2: نمط ديناميكي (يلتقط أي tikmatrixXXX جديد)
        if TIKMATRIX_FALLBACK_PATTERN.match(actual) and actual != requested_lower:
            return False, actual
        # تحقق 3: عدم تطابق الجوهري (ليس مطابقاً للمطلوب)
        if actual != requested_lower and 'tikmatrix' in actual:
            return False, actual
    return True, None


def extract_bio_v210(content):
    if not content:
        return None
    for pattern in BIO_PATTERNS:
        m = re.search(pattern, content, re.DOTALL)
        if m:
            bio = m.group(1).strip()
            if bio and 2 < len(bio) < 500:
                if 'no about' in bio.lower() or bio.startswith('[') or bio.startswith('http'):
                    continue
                return bio
    return None


def extract_fields(content):
    if not content:
        return {}
    data = {}
    m = re.search(PATTERNS['country_flag'], content)
    if m:
        data['country'] = re.sub(r'\s+', ' ', m.group(1)).strip()
        data['country_source'] = 'flag_emoji'
    else:
        m = re.search(PATTERNS['country_globe'], content)
        if m:
            c = re.sub(r'\s+', ' ', m.group(1)).strip()
            if 2 < len(c) < 50:
                data['country'] = c
                data['country_source'] = 'globe_emoji'
    for key in ['language', 'user_id', 'sec_uid', 'created', 'nickname', 'avatar']:
        m = re.search(PATTERNS[key], content)
        if m:
            data[key] = m.group(1).strip()
    for key in ['followers', 'following', 'hearts', 'videos', 'friends']:
        m = re.search(PATTERNS[key], content)
        if m:
            value = (m.group(1) or m.group(2) or '').replace(',', '').strip()
            if value.isdigit():
                data[key] = int(value)
    bio = extract_bio_v210(content)
    if bio:
        data['bio'] = bio
    return data


def correct_country(username, data):
    original = data.get('country')
    language = (data.get('language') or '').lower()[:2]
    bio = data.get('bio', '') or ''
    nickname = data.get('nickname', '') or ''
    username_lower = username.lower()
    log = []

    if username_lower in LOCAL_VERIFIED_DB:
        ver = LOCAL_VERIFIED_DB[username_lower]
        log.append(f"تحقق محلي معتمد")
        return _build(ver['country'], 'local_verified', original, log, 99, ver.get('region'))

    if username_lower in CELEBRITIES:
        celeb = CELEBRITIES[username_lower]
        log.append(f"قاعدة المشاهير: {celeb['name']}")
        return _build(celeb['country'], 'celebrity_database', original, log, 98, celeb.get('region'))

    for misleading, real in MISLEADING_PATTERNS.items():
        if misleading in username_lower:
            if original != real:
                log.append(f"اسم مضلل: {misleading}")
                return _build(real, 'misleading_name_fix', original, log, 85)

    for keyword, c in USERNAME_KEYWORDS_SORTED:
        if keyword in username_lower:
            if not original or original != c:
                log.append(f"كلمة في اسم المستخدم: {keyword}")
                return _build(c, 'username_keyword', original, log, 95 if not original else 88)
            else:
                log.append(f"تأكيد اسم المستخدم")
                return _build(c, 'username_confirmed', original, log, 96)

    for flag, c in FLAG_EMOJI_TO_COUNTRY.items():
        if flag in bio or flag in nickname:
            if c != original:
                log.append(f"علم في الوصف: {flag}")
                return _build(c, 'bio_flag', original, log, 92)

    for pattern, c, name in BIO_SCRIPT_TO_COUNTRY:
        if re.search(pattern, bio + nickname):
            # ✅ تطوير v2.1.1 - تجاوز إذا c=None (مثل العربية)
            if c is None:
                log.append(f"نص {name} - لا يحدد دولة")
                continue
            if original != c:
                log.append(f"نص: {name}")
                return _build(c, 'bio_script', original, log, 88)

    if language and language in LANGUAGE_TO_COUNTRY and language not in GLOBAL_LANGUAGES:
        primary, valid = LANGUAGE_TO_COUNTRY[language]
        # ✅ إصلاح v2.1.1 - تجاوز إذا primary=None (مثل العربية)
        if primary is None:
            # اللغة تغطي عدة دول - لا نخمن، نحتفظ ببيانات TikMatrix
            if original and original in valid:
                log.append(f"اللغة {language} توافق بيانات TikMatrix")
            # لا return - نترك المنطق يكمل للطبقات التالية
        else:
            if original and original not in valid and original not in TRUSTED_TIKMATRIX_COUNTRIES:
                log.append(f"تصحيح اللغة: {language}")
                return _build(primary, 'language_override', original, log, 72)
            if original in TIKTOK_SERVER_COUNTRIES and language != 'en':
                log.append(f"فلتر سيرفر TikTok")
                return _build(primary, 'tiktok_server_filter', original, log, 70)

    for tld, c in TLD_TO_COUNTRY.items():
        if username_lower.endswith(tld):
            if tld in AMBIGUOUS_TLDS:
                if original and original != c:
                    log.append(f"TLD غامض")
                    result = _build(original, 'ambiguous_tld', original, log, 50, None)
                    result['ambiguous'] = True
                    result['alternative_country'] = c
                    return result
            elif original in EXPANDED_SUSPICIOUS or (original and original != c):
                log.append(f"نطاق الدولة: {tld}")
                return _build(c, 'tld_priority_fix', original, log, 88)

    if original in EXPANDED_SUSPICIOUS and language in LANGUAGE_TO_COUNTRY:
        primary, valid = LANGUAGE_TO_COUNTRY[language]
        # ✅ إصلاح v2.1.1 - تجاوز إذا primary=None
        if primary is None:
            log.append(f"اللغة {language} غامضة - تخطي الفلتر")
        if original not in valid:
            log.append(f"دولة مشبوهة")
            return _build(primary, 'suspicious_filter', original, log, 65)

    # ✅ إصلاح v2.1.1 - ضمان return صالح حتى عند انعدام البيانات
    if original:
        log.append(f"بيانات TikMatrix")
        return _build(original, data.get('country_source', 'tikmatrix'), original, log, 75)
    
    # حالة original = None - إرجاع كائن صالح بدلاً من None ضمنياً
    log.append("لا توجد بيانات دولة - إرجاع غير محدد")
    return _build('غير محدد', 'no_data', None, log, 0, None)


def _build(country, source, original, log, confidence, region=None):
    flag = ''
    for emoji, c in FLAG_EMOJI_TO_COUNTRY.items():
        if c == country:
            flag = emoji
            break
    return {
        'country': country, 'flag': flag, 'confidence': confidence,
        'source': source, 'original_tikmatrix': original, 'corrections': log,
        'preset_region': region, 'ambiguous': False, 'alternative_country': None,
    }


def detect_expatriate(nationality, residence, source):
    # ✅ إصلاح v2.1.1 - تحقق صارم من الإقامة الفعلية
    if not nationality or not residence:
        return {'is_expat': False, 'confidence': 0}
    if nationality == residence:
        return {'is_expat': False, 'confidence': 0}
    # تجنب الإعلان الخاطئ للمغترب عند البيانات الخام الفارغة
    if not nationality.strip() or not residence.strip():
        return {'is_expat': False, 'confidence': 0}
    if nationality.strip() == 'غير محدد' or residence.strip() == 'غير محدد':
        return {'is_expat': False, 'confidence': 0}
    conf_map = {'local_verified': 99, 'celebrity_database': 95, 'bio_flag': 90,
                'bio_script': 85, 'tld_domain': 88, 'username_keyword': 80}
    return {'is_expat': True, 'confidence': conf_map.get(source, 60)}


def get_smart_region(country, residence, bio, nickname, username, preset_region=None):
    """
    ✅ v2.1.2 Emergency Fix - حذف نهائي للعاصمة الافتراضية
    لا تخمين الرياض للسعوديين أو أي عاصمة لأي جنسية
    الصدق أولاً: إرجاع None عند عدم وجود دليل فعلي
    """
    # الطبقة 1: حساب موثق يدوياً (99%)
    if preset_region:
        return {'region_ar': preset_region, 'region_en': preset_region,
                'confidence': 99, 'source': 'preset', 'is_estimate': False}
    
    actual_location = residence if residence else country
    text = f"{bio or ''} {nickname or ''} {username or ''}".strip()
    
    # الطبقة 2: بحث فعلي في BIO/Nickname/Username (90%)
    if text and actual_location:
        r = lookup_region(actual_location, text)
        if r:
            r['source'] = 'bio_extraction'
            r['is_estimate'] = False
            return r
    
    # الطبقة 3: بحث بدولة الجنسية إذا اختلفت (85%)
    if text and country and country != actual_location:
        r = lookup_region(country, text)
        if r:
            r['source'] = 'bio_nationality'
            r['confidence'] = 85
            r['is_estimate'] = False
            return r
    
    # ❌ محذوف نهائياً v2.1.2:
    # - lookup_capital(actual_location) للإقامة
    # - lookup_capital(country) للجنسية
    # السبب: المستخدم رفض تخمين العاصمة لأنه غير علمي
    
    # الصدق: لا نجد دليل = إرجاع None
    return None


def _raw_from_tikwm(tikwm_json, region_iso):
    """✅ v2.1.6 - تحويل JSON من tikwm إلى بنية raw المتوقعة في النظام"""
    data = (tikwm_json or {}).get('data') or {}
    user = data.get('user') or {}
    stats = data.get('stats') or {}
    bio = user.get('signature') or None
    nickname = user.get('nickname') or None
    created_ts = user.get('createTime')
    try:
        created_str = datetime.fromtimestamp(int(created_ts)).strftime('%Y-%m-%d') if created_ts else None
    except Exception:
        created_str = None
    # استنتاج رمز اللغة من bio+nickname (بلا تخمين إذا لم توجد أغلبية واضحة)
    lang = _detect_language_from_text(((bio or '') + ' ' + (nickname or '')))
    # تحويل region ISO إلى اسم دولة إنجليزي
    country = REGION_ISO_TO_COUNTRY.get((region_iso or '').upper()) if region_iso else None
    return {
        'country': country,  # قد تكون None إذا لم يوجد region (لا تخمين)
        'language': lang,
        'user_id': str(user.get('id')) if user.get('id') else None,
        'sec_uid': user.get('secUid'),
        'created': created_str,
        'nickname': nickname,
        'avatar': user.get('avatarLarger') or user.get('avatarMedium') or user.get('avatarThumb'),
        'bio': bio,
        'followers': int(stats.get('followerCount') or 0),
        'following': int(stats.get('followingCount') or 0),
        'hearts': int(stats.get('heartCount') or 0),
        'videos': int(stats.get('videoCount') or 0),
        'friends': 0,
        'verified': bool(user.get('verified')),
        'bio_link': ((user.get('bioLink') or {}).get('link')),
    }


def lookup_user(username):
    username = username.strip().lower().lstrip('@')
    username = re.sub(r'https?://(?:www\.)?tiktok\.com/@?', '', username)
    username = username.split('?')[0].split('/')[0]

    fetch = fetch_user(username)
    if not fetch['success']:
        return {'success': False, 'username': username,
                'error': 'تعذّر جلب بيانات الحساب من جميع المصادر (tikwm + jina + tikmatrix). ربما الحساب غير موجود أو خاص أو المصدر تحت ضغط.'}

    # ✅ v2.1.6 - مسار tikwm الأساسي
    if fetch.get('tikwm_json'):
        raw = _raw_from_tikwm(fetch['tikwm_json'], fetch.get('region_iso'))
        # إذا لدينا region رسمي (من فيديوهات tikwm) نبني correction مباشرة بثقة 95%
        if raw.get('country'):
            country = raw['country']
            flag = COUNTRY_TO_FLAG_EMOJI.get(country, '')
            region_info = get_smart_region(
                country, None,
                raw.get('bio'), raw.get('nickname'), username,
                None,
            )
            expat = detect_expatriate(country, None, 'tikwm_official')
            return {
                'success': True, 'username': username,
                'nickname': raw.get('nickname'),
                'country': country, 'country_flag': flag,
                'country_source': 'tikwm_official_region',
                'country_original': country,
                'language': raw.get('language'),
                'followers': raw.get('followers', 0), 'following': raw.get('following', 0),
                'hearts': raw.get('hearts', 0), 'videos': raw.get('videos', 0),
                'friends': raw.get('friends', 0),
                'user_id': raw.get('user_id'), 'sec_uid': raw.get('sec_uid'),
                'created': raw.get('created'), 'bio': raw.get('bio'),
                'avatar': raw.get('avatar'),
                'confidence': 95,
                'corrections_log': [f"region رسمي من tikwm/user/posts: {fetch.get('region_iso')}"],
                'proxy_used': 'tikwm', 'fetch_time': fetch['time'],
                'is_expatriate': expat['is_expat'],
                'expat_confidence': expat['confidence'],
                'region_info': region_info,
                'ambiguous': False,
                'alternative_country': None,
                'verified': raw.get('verified', False),
                'bio_link': raw.get('bio_link'),
            }
        # tikwm نجح لكن بدون region → Multi-Signal على bio/nickname/username فقط
        correction = correct_country(username, raw)
        if not correction or not isinstance(correction, dict):
            return {'success': False, 'username': username,
                    'error': 'فشل تحديد الجنسية', 'reason': 'correction_failed'}
        raw_country = raw.get('country')
        corrected_country = correction.get('country')
        actual_residence = raw_country if (raw_country and raw_country != corrected_country) else None
        region_info = get_smart_region(
            corrected_country, actual_residence,
            raw.get('bio'), raw.get('nickname'), username,
            correction.get('preset_region'),
        )
        expat = detect_expatriate(corrected_country, actual_residence, correction.get('source', 'unknown'))
        return {
            'success': True, 'username': username,
            'nickname': raw.get('nickname'),
            'country': correction.get('country'), 'country_flag': correction.get('flag', ''),
            'country_source': correction.get('source', 'multi_signal_no_region'),
            'country_original': correction.get('original_tikmatrix'),
            'language': raw.get('language'),
            'followers': raw.get('followers', 0), 'following': raw.get('following', 0),
            'hearts': raw.get('hearts', 0), 'videos': raw.get('videos', 0),
            'friends': raw.get('friends', 0),
            'user_id': raw.get('user_id'), 'sec_uid': raw.get('sec_uid'),
            'created': raw.get('created'), 'bio': raw.get('bio'),
            'avatar': raw.get('avatar'),
            'confidence': min(correction.get('confidence', 50), 75),  # سقف 75% بلا region رسمي
            'corrections_log': correction.get('corrections', []) + ['لا يوجد region رسمي - Multi-Signal فقط'],
            'proxy_used': 'tikwm+multi_signal', 'fetch_time': fetch['time'],
            'is_expatriate': expat['is_expat'],
            'expat_confidence': expat['confidence'],
            'region_info': region_info,
            'ambiguous': correction.get('ambiguous', False),
            'alternative_country': correction.get('alternative_country'),
            'verified': raw.get('verified', False),
            'bio_link': raw.get('bio_link'),
        }

    # ✅ v2.1.6 - مسار jina/tikmatrix الاحتياطي (السلوك الأصلي)
    is_match, actual = verify_username_match(fetch.get('content') or '', username)
    if not is_match:
        return {'success': False, 'username': username,
                'error': f'الحساب @{username} غير موجود أو خاص',
                'reason': 'account_not_found'}

    raw = extract_fields(fetch.get('content') or '')
    correction = correct_country(username, raw)
    
    # ✅ إصلاح v2.1.1 - التحقق من صحة correction
    if not correction or not isinstance(correction, dict):
        return {'success': False, 'username': username,
                'error': 'فشل تصحيح الجنسية',
                'reason': 'correction_failed'}
    
    # ✅ إصلاح v2.1.1 - الإقامة الفعلية = بيانات TikMatrix الخام
    raw_country = raw.get('country')
    corrected_country = correction.get('country')
    
    # تحديد الإقامة الفعلية - فقط إذا اختلفت عن الجنسية المصححة
    actual_residence = raw_country if (raw_country and raw_country != corrected_country) else None
    
    region_info = get_smart_region(
        corrected_country, actual_residence,
        raw.get('bio'), raw.get('nickname'), username,
        correction.get('preset_region'),
    )
    
    # ✅ إصلاح v2.1.1 - استخدام الإقامة الفعلية فقط
    expat = detect_expatriate(corrected_country, actual_residence, correction.get('source', 'unknown'))

    return {
        'success': True, 'username': username,
        'nickname': raw.get('nickname'),
        'country': correction['country'], 'country_flag': correction['flag'],
        'country_source': correction['source'],
        'country_original': correction['original_tikmatrix'],
        'language': raw.get('language'),
        'followers': raw.get('followers', 0), 'following': raw.get('following', 0),
        'hearts': raw.get('hearts', 0), 'videos': raw.get('videos', 0),
        'friends': raw.get('friends', 0),
        'user_id': raw.get('user_id'), 'sec_uid': raw.get('sec_uid'),
        'created': raw.get('created'), 'bio': raw.get('bio'),
        'avatar': raw.get('avatar'),
        'confidence': correction['confidence'],
        'corrections_log': correction['corrections'],
        'proxy_used': fetch.get('proxy', '-'), 'fetch_time': fetch.get('time', 0),
        'is_expatriate': expat['is_expat'],
        'expat_confidence': expat['confidence'],
        'region_info': region_info,
        'ambiguous': correction.get('ambiguous', False),
        'alternative_country': correction.get('alternative_country'),
        # ✅ v2.1.7-Light-Fix3.1 - تمرير حقول الإقامة الفعلية للواجهة
        'actual_residence': fetch.get('actual_residence'),
        'previous_residence': fetch.get('previous_residence'),
        'residence_confidence': fetch.get('residence_confidence', 0),
        'residence_type': fetch.get('residence_type', '—'),
        'videos_analyzed': fetch.get('videos_count', 0),
        'region_distribution': fetch.get('region_distribution') or {},
        'timezone_match': fetch.get('timezone_match'),
    }


# ═══════════════════════════════════════════════════════════════
# 🌍 ترجمة الدول
# ═══════════════════════════════════════════════════════════════
# ✅ v2.1.7 - ترجمة عربية شاملة تغطي جميع دول ISO 3166 في REGION_ISO_TO_COUNTRY
COUNTRY_AR = {
    # الدول العربية
    'Saudi Arabia': 'المملكة العربية السعودية', 'United Arab Emirates': 'الإمارات',
    'Egypt': 'مصر', 'Kuwait': 'الكويت', 'Qatar': 'قطر', 'Bahrain': 'البحرين',
    'Oman': 'عُمان', 'Jordan': 'الأردن', 'Lebanon': 'لبنان', 'Iraq': 'العراق',
    'Yemen': 'اليمن', 'Palestine': 'فلسطين', 'Morocco': 'المغرب',
    'Algeria': 'الجزائر', 'Tunisia': 'تونس', 'Libya': 'ليبيا', 'Sudan': 'السودان',
    'Somalia': 'الصومال', 'Mauritania': 'موريتانيا', 'Djibouti': 'جيبوتي',
    'Comoros': 'جزر القمر',
    # أوروبا وأمريكا الشمالية وأستراليا
    'United States': 'الولايات المتحدة', 'United Kingdom': 'المملكة المتحدة',
    'Canada': 'كندا', 'Australia': 'أستراليا', 'New Zealand': 'نيوزيلندا',
    'Ireland': 'أيرلندا',
    'France': 'فرنسا', 'Germany': 'ألمانيا', 'Italy': 'إيطاليا',
    'Spain': 'إسبانيا', 'Portugal': 'البرتغال', 'Netherlands': 'هولندا',
    'Belgium': 'بلجيكا', 'Switzerland': 'سويسرا', 'Austria': 'النمسا',
    'Sweden': 'السويد', 'Norway': 'النرويج', 'Finland': 'فنلندا',
    'Denmark': 'الدنمارك', 'Poland': 'بولندا',
    'Czech Republic': 'تشيكيا', 'Slovakia': 'سلوفاكيا',
    'Hungary': 'المجر', 'Romania': 'رومانيا', 'Bulgaria': 'بلغاريا',
    'Greece': 'اليونان', 'Croatia': 'كرواتيا', 'Slovenia': 'سلوفينيا',
    'Serbia': 'صربيا', 'Bosnia and Herzegovina': 'البوسنة والهرسك',
    'North Macedonia': 'مقدونيا الشمالية', 'Albania': 'ألبانيا',
    'Montenegro': 'الجبل الأسود', 'Kosovo': 'كوسوفو',
    'Estonia': 'إستونيا', 'Latvia': 'لاتفيا', 'Lithuania': 'ليتوانيا',
    'Iceland': 'أيسلندا', 'Luxembourg': 'لوكسمبورغ', 'Malta': 'مالطا',
    'Cyprus': 'قبرص', 'Monaco': 'موناكو', 'Andorra': 'أندورا',
    'San Marino': 'سان مارينو', 'Vatican City': 'الفاتيكان',
    'Liechtenstein': 'ليختنشتاين',
    # روسيا ودول السوفيات سابقاً
    'Russia': 'روسيا', 'Ukraine': 'أوكرانيا', 'Belarus': 'روسيا البيضاء',
    'Moldova': 'مولدوفا', 'Georgia': 'جورجيا', 'Armenia': 'أرمينيا',
    'Azerbaijan': 'أذربيجان', 'Kazakhstan': 'كازاخستان',
    'Uzbekistan': 'أوزبكستان', 'Kyrgyzstan': 'قيرغيزستان',
    'Tajikistan': 'طاجيكستان', 'Turkmenistan': 'تركمانستان',
    # الشرق الأوسط الأوسع
    'Turkey': 'تركيا', 'Israel': 'إسرائيل', 'Iran': 'إيران',
    # جنوب آسيا
    'India': 'الهند', 'Pakistan': 'باكستان', 'Bangladesh': 'بنغلاديش',
    'Sri Lanka': 'سريلانكا', 'Nepal': 'نيبال', 'Afghanistan': 'أفغانستان',
    'Bhutan': 'بوتان', 'Maldives': 'جزر المالديف',
    # شرق آسيا
    'China': 'الصين', 'Japan': 'اليابان', 'South Korea': 'كوريا الجنوبية',
    'North Korea': 'كوريا الشمالية', 'Taiwan': 'تايوان',
    'Hong Kong': 'هونغ كونغ', 'Macau': 'ماكاو', 'Mongolia': 'منغوليا',
    # جنوب شرق آسيا
    'Singapore': 'سنغافورة', 'Malaysia': 'ماليزيا',
    'Indonesia': 'إندونيسيا', 'Thailand': 'تايلاند',
    'Vietnam': 'فيتنام', 'Philippines': 'الفلبين',
    'Myanmar': 'ميانمار', 'Cambodia': 'كمبوديا', 'Laos': 'لاوس',
    'Brunei': 'بروناي', 'Timor-Leste': 'تيمور الشرقية',
    # أمريكا اللاتينية والكاريبي
    'Brazil': 'البرازيل', 'Mexico': 'المكسيك',
    'Argentina': 'الأرجنتين', 'Chile': 'تشيلي',
    'Colombia': 'كولومبيا', 'Peru': 'البيرو',
    'Venezuela': 'فنزويلا', 'Uruguay': 'الأوروغواي',
    'Paraguay': 'الباراغواي', 'Bolivia': 'بوليفيا',
    'Ecuador': 'الإكوادور', 'Guyana': 'غيانا', 'Suriname': 'سورينام',
    'Guatemala': 'غواتيمالا', 'Honduras': 'هندوراس',
    'El Salvador': 'السلفادور', 'Nicaragua': 'نيكاراغوا',
    'Costa Rica': 'كوستاريكا', 'Panama': 'بنما',
    'Dominican Republic': 'جمهورية الدومينيكان',
    'Cuba': 'كوبا', 'Haiti': 'هايتي', 'Jamaica': 'جامايكا',
    'Puerto Rico': 'بورتو ريكو', 'Trinidad and Tobago': 'ترينيداد وتوباغو',
    'Bahamas': 'الباهاما', 'Barbados': 'بربادوس', 'Belize': 'بليز',
    # أفريقيا جنوب الصحراء
    'Nigeria': 'نيجيريا', 'Kenya': 'كينيا', 'Ethiopia': 'إثيوبيا',
    'Ghana': 'غانا', 'South Africa': 'جنوب أفريقيا',
    'Tanzania': 'تنزانيا', 'Uganda': 'أوغندا',
    'Ivory Coast': 'ساحل العاج', 'Senegal': 'السنغال',
    'Cameroon': 'الكاميرون', 'Mali': 'مالي',
    'Burkina Faso': 'بوركينا فاسو', 'Niger': 'النيجر', 'Chad': 'تشاد',
    'Angola': 'أنغولا', 'Mozambique': 'موزمبيق',
    'Zimbabwe': 'زيمبابوي', 'Zambia': 'زامبيا', 'Malawi': 'مالاوي',
    'Botswana': 'بوتسوانا', 'Namibia': 'ناميبيا',
    'Lesotho': 'ليسوتو', 'Eswatini': 'إسواتيني',
    'Madagascar': 'مدغشقر', 'Mauritius': 'موريشيوس',
    'Seychelles': 'سيشل', 'Rwanda': 'رواندا', 'Burundi': 'بوروندي',
    'DR Congo': 'جمهورية الكونغو الديمقراطية',
    'Congo': 'الكونغو', 'Gabon': 'الغابون',
    'Equatorial Guinea': 'غينيا الاستوائية',
    'Central African Republic': 'جمهورية أفريقيا الوسطى',
    'South Sudan': 'جنوب السودان', 'Eritrea': 'إريتريا',
    'Liberia': 'ليبيريا', 'Sierra Leone': 'سيراليون',
    'Guinea': 'غينيا', 'Gambia': 'غامبيا',
    'Guinea-Bissau': 'غينيا بيساو', 'Cape Verde': 'الرأس الأخضر',
    'Togo': 'توغو', 'Benin': 'بنين',
    'Sao Tome and Principe': 'ساو تومي وبرينسيبي',
    # أوقيانوسيا
    'Fiji': 'فيجي', 'Papua New Guinea': 'بابوا غينيا الجديدة',
    'Solomon Islands': 'جزر سليمان', 'Vanuatu': 'فانواتو',
    'Samoa': 'ساموا', 'Tonga': 'تونغا',
    'Kiribati': 'كيريباتي', 'Tuvalu': 'توفالو', 'Nauru': 'ناورو',
    'Palau': 'بالاو', 'Micronesia': 'ميكرونيزيا',
    'Marshall Islands': 'جزر مارشال',
}


# ═══════════════════════════════════════════════════════════════
# 🎨 CSS احترافي + Responsive
# ═══════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+Arabic:wght@400;500;700;900&family=Tajawal:wght@400;700;900&display=swap');

* { font-family: 'Noto Sans Arabic', 'Tajawal', sans-serif !important; }
html, body, [class*="css"] { direction: rtl; text-align: right; }

/* إخفاء عناصر Streamlit */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

.main-title {
    font-size: 3.5rem; font-weight: 900; text-align: center;
    background: linear-gradient(135deg, #F59E0B, #FCD34D, #F59E0B);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    margin: 1.5rem 0 0.5rem 0;
    letter-spacing: -1px;
}
.subtitle {
    text-align: center; color: #94A3B8; font-size: 1rem;
    margin-bottom: 2rem; font-weight: 500;
}

/* البطاقات الرئيسية */
.country-card {
    background: linear-gradient(135deg, #1E3A8A, #3B82F6);
    border-radius: 20px; padding: 1.5rem; text-align: center;
    box-shadow: 0 10px 40px rgba(59, 130, 246, 0.4); margin: 1rem 0;
    border: 1px solid rgba(255, 255, 255, 0.1);
}
.residence-card {
    background: linear-gradient(135deg, #065F46, #10B981);
    border-radius: 20px; padding: 1.5rem; text-align: center;
    box-shadow: 0 10px 40px rgba(16, 185, 129, 0.4); margin: 1rem 0;
    border: 1px solid rgba(255, 255, 255, 0.1);
}
.region-card {
    background: linear-gradient(135deg, #7C2D12, #F97316);
    border-radius: 20px; padding: 1.5rem; text-align: center;
    box-shadow: 0 10px 40px rgba(249, 115, 22, 0.4); margin: 1rem 0;
    border: 1px solid rgba(255, 255, 255, 0.1);
}
.region-card-estimate {
    background: linear-gradient(135deg, #713F12, #CA8A04);
    border-radius: 20px; padding: 1.5rem; text-align: center;
    box-shadow: 0 10px 40px rgba(202, 138, 4, 0.3); margin: 1rem 0;
    border: 2px dashed #FBBF24;
}
.alt-country-card {
    background: linear-gradient(135deg, #7C2D12, #C2410C);
    border-radius: 20px; padding: 1.5rem; text-align: center;
    box-shadow: 0 8px 32px rgba(194, 65, 12, 0.3); margin: 1rem 0;
}

.country-flag { font-size: 3.5rem; line-height: 1; margin-bottom: 0.6rem; }
.country-name { color: #F1F5F9; font-size: 1.5rem; font-weight: 700; }

.confidence-high { background: linear-gradient(135deg, #10B981, #059669); color: white; padding: 0.5rem 1.2rem; border-radius: 999px; font-weight: 700; box-shadow: 0 4px 16px rgba(16,185,129,0.3); }
.confidence-medium { background: linear-gradient(135deg, #F59E0B, #D97706); color: white; padding: 0.5rem 1.2rem; border-radius: 999px; font-weight: 700; box-shadow: 0 4px 16px rgba(245,158,11,0.3); }
.confidence-low { background: linear-gradient(135deg, #EF4444, #DC2626); color: white; padding: 0.5rem 1.2rem; border-radius: 999px; font-weight: 700; box-shadow: 0 4px 16px rgba(239,68,68,0.3); }

.expat-badge {
    background: linear-gradient(135deg, #7C3AED, #A78BFA); color: white;
    padding: 0.5rem 1.3rem; border-radius: 999px; font-weight: 700;
    display: inline-block; margin: 0.5rem 0;
    box-shadow: 0 4px 16px rgba(124, 58, 237, 0.3);
}
.estimate-badge {
    background: #FBBF24; color: #0F172A; padding: 0.4rem 1rem;
    border-radius: 999px; font-weight: 700; display: inline-block;
    margin-top: 0.8rem; font-size: 0.8rem;
}
.ambiguous-badge {
    background: linear-gradient(135deg, #CA8A04, #FBBF24); color: #0F172A;
    padding: 0.5rem 1.2rem; border-radius: 999px; font-weight: 700;
    display: inline-block; margin: 0.5rem 0;
}

/* بطاقة معلومات الحساب - ✅ v2.1.5 تحسين RTL ومسافات */
.result-card {
    background: linear-gradient(135deg, rgba(15, 23, 42, 0.8), rgba(30, 41, 59, 0.6));
    border: 1px solid rgba(245, 158, 11, 0.2);
    border-radius: 20px; padding: 2rem 2.2rem; margin: 1.2rem 0;
    backdrop-filter: blur(10px);
    direction: rtl; text-align: right;
    font-family: 'Noto Sans Arabic', 'Tajawal', sans-serif;
    line-height: 1.85;
}
.result-card * { font-family: 'Noto Sans Arabic', 'Tajawal', sans-serif; }

/* الإحصائيات - ✅ v2.1.5 padding أوسع */
.stat-card {
    background: linear-gradient(135deg, rgba(15, 23, 42, 0.8), rgba(30, 41, 59, 0.6));
    border: 1px solid rgba(245, 158, 11, 0.15);
    border-radius: 16px; padding: 1.4rem 1rem; text-align: center;
    transition: transform 0.2s, box-shadow 0.2s;
    direction: rtl;
    font-family: 'Noto Sans Arabic', 'Tajawal', sans-serif;
    margin: 0.4rem 0;
}
.stat-card:hover {
    transform: translateY(-4px);
    box-shadow: 0 8px 24px rgba(245, 158, 11, 0.2);
}
.stat-number { font-size: 1.8rem; font-weight: 900; color: #F59E0B; }
.stat-label { color: #94A3B8; font-size: 0.85rem; margin-top: 0.3rem; }

/* تفاصيل تقنية - ✅ v2.1.5 RTL مضمون */
.tech-details-card {
    background: linear-gradient(135deg, rgba(15, 23, 42, 0.95), rgba(30, 41, 59, 0.8));
    border: 1px solid rgba(245, 158, 11, 0.3);
    border-radius: 16px; padding: 1.6rem 1.8rem; margin: 1.2rem 0;
    direction: rtl; text-align: right;
    font-family: 'Noto Sans Arabic', 'Tajawal', sans-serif;
}
.tech-grid {
    display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 1rem; margin-top: 1rem;
}
.tech-item {
    background: rgba(15, 23, 42, 0.6);
    border-right: 3px solid #F59E0B;
    border-radius: 8px; padding: 1rem;
}
.tech-label {
    color: #94A3B8; font-size: 0.8rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: 1px; margin-bottom: 0.4rem;
}
.tech-value {
    color: #F1F5F9; font-family: 'Courier New', monospace;
    font-size: 0.95rem; word-break: break-all;
}

/* جدول البحث الجماعي */
.bulk-table {
    background: rgba(15, 23, 42, 0.6);
    border-radius: 16px; padding: 0;
    overflow: hidden; margin: 1rem 0;
}

/* تصميم Responsive للجوال */
@media (max-width: 768px) {
    .main-title { font-size: 2.2rem; }
    .subtitle { font-size: 0.85rem; }
    .country-flag { font-size: 2.5rem; }
    .country-name { font-size: 1.1rem; }
    .stat-number { font-size: 1.3rem; }
    .stat-label { font-size: 0.7rem; }
    .result-card { padding: 1rem; }
    .tech-grid { grid-template-columns: 1fr; }
    .country-card, .residence-card, .region-card { padding: 1rem; }
}
@media (max-width: 480px) {
    .main-title { font-size: 1.8rem; }
    .country-flag { font-size: 2rem; }
    .stat-card { padding: 0.7rem; }
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] { gap: 1rem; }
.stTabs [data-baseweb="tab"] {
    background: rgba(15, 23, 42, 0.6); border-radius: 12px;
    padding: 0.5rem 1.5rem; color: #94A3B8; font-weight: 700;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #F59E0B, #FCD34D);
    color: #0F172A;
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# 🎨 العنوان
# ═══════════════════════════════════════════════════════════════
st.markdown('<div class="main-title">🦅 بَصِير</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle" dir="rtl">مولّد ذكي لمعلومات حسابات TikTok</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# 🎨 دوال العرض
# ═══════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════
# 🗺️ ✅ v2.1.7-Light-Fix3.2-Map - خريطة Choropleth دول فقط
# ═══════════════════════════════════════════════════════════════
def render_country_choropleth(region_distribution, actual_residence=None):
    """Fix3.2-Map: خريطة عالمية تظلّل دول الفيديوهات فقط
    القيود: لا POI, لا مدن, لا Geocoding خارجي, GeoJSON محلّي فقط
    """
    if not region_distribution:
        return None

    try:
        import folium
        from streamlit_folium import st_folium
        import json as _json
        import os as _os
    except Exception:
        return None

    # ✅ Fix3.2-Map-Patch: ابحث في data/ أوّلاً ثمّ في الجذر (توافق GitHub)
    _base = _os.path.dirname(__file__)
    _candidates = [
        _os.path.join(_base, 'data', 'world_countries.geo.json'),
        _os.path.join(_base, 'world_countries.geo.json'),
        'data/world_countries.geo.json',
        'world_countries.geo.json',
    ]
    geojson_path = None
    for _c in _candidates:
        if _os.path.exists(_c):
            geojson_path = _c
            break
    if geojson_path is None:
        return None

    try:
        with open(geojson_path, 'r', encoding='utf-8') as _f:
            world_geo = _json.load(_f)
    except Exception:
        return None

    # تحويل ISO-2 من region_distribution إلى صفّ موحّد للخريطة
    total = sum(region_distribution.values())
    rows = []
    for iso2, cnt in region_distribution.items():
        pct = round(cnt * 100 / total, 1) if total else 0
        rows.append({'iso2': iso2.upper(), 'count': cnt, 'pct': pct})

    # إنشاء dict سريع للوصول من ISO-2
    dist_by_iso = {r['iso2']: r for r in rows}

    # مركز الخريطة = أول دولة بأكبر عدد، fallback (20, 0)
    center_lat, center_lon = 20.0, 0.0

    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=2,
        tiles='CartoDB positron',
        prefer_canvas=True,
    )

    # دالة لون التظليل بناءً على العدد
    def _style_function(feature):
        iso2 = (feature.get('properties', {}).get('ISO3166-1-Alpha-2') or '').upper()
        if iso2 in dist_by_iso:
            cnt = dist_by_iso[iso2]['count']
            # تدرّج لوني #F59E0B
            if cnt >= 4:
                fill = '#D97706'  # غامق
                opacity = 0.85
            elif cnt >= 2:
                fill = '#F59E0B'  # متوسط
                opacity = 0.70
            else:
                fill = '#FCD34D'  # فاتح
                opacity = 0.55
            return {
                'fillColor': fill,
                'color': '#0F172A',
                'weight': 1.5,
                'fillOpacity': opacity,
            }
        # دول غير مذكورة في التوزيع
        return {
            'fillColor': '#F1F5F9',
            'color': '#CBD5E1',
            'weight': 0.3,
            'fillOpacity': 0.1,
        }

    # إضافة طبقة الخريطة
    folium.GeoJson(
        world_geo,
        name='dist',
        style_function=_style_function,
        highlight_function=lambda x: {'weight': 3, 'color': '#F59E0B'},
        tooltip=folium.GeoJsonTooltip(
            fields=['name', 'ISO3166-1-Alpha-2'],
            aliases=['الدولة:', 'الرمز:'],
            localize=True,
            sticky=False,
            labels=True,
            style="""
                background-color: #0F172A;
                color: #F1F5F9;
                font-family: 'Noto Sans Arabic', Tajawal, sans-serif;
                font-size: 13px;
                padding: 6px;
                border-radius: 6px;
                direction: rtl;
            """,
        ),
    ).add_to(m)

    return m


def display_single_result(result):
    """عرض نتيجة حساب واحد بتصميم احترافي - ✅ v2.1.1"""
    if not result.get('success'):
        st.error(f"❌ {result.get('error', 'فشل البحث')}")
        if result.get('reason') == 'account_not_found':
            st.info("ℹ️ **الأسباب المحتملة**: حساب خاص، محذوف، محظور، أو خطأ إملائي")
        return
    
    # ✅ تطوير #5 v2.1.1 - عرض الأفاتار + الاسم بصورة بارزة
    avatar = result.get('avatar')
    nickname = result.get('nickname', 'غير معروف')
    username_display = result.get('username', '')
    
    if avatar:
        st.markdown(f'''
        <div dir="rtl" style="
            background:#0F172A; padding:16px; border-radius:12px;
            display:flex; align-items:center; gap:16px; margin-bottom:12px;
            border-right:4px solid #F59E0B;
            font-family:'Noto Sans Arabic','Tajawal',sans-serif;
        ">
            <img src="{avatar}" style="width:80px;height:80px;border-radius:50%;border:3px solid #F59E0B;object-fit:cover;" />
            <div style="color:#F1F5F9;">
                <div style="font-size:1.3rem;font-weight:700;">{nickname}</div>
                <div style="color:#93C5FD;font-size:0.95rem;">@{username_display}</div>
            </div>
        </div>
        ''', unsafe_allow_html=True)
    
    # ✅ تطوير #6 v2.1.1 - إشعار للحسابات الجديدة (أقل من 6 أشهر)
    created = result.get('created', '')
    if created:
        try:
            from datetime import datetime
            for fmt in ['%m/%d/%Y %I:%M:%S %p', '%Y-%m-%d', '%m/%d/%Y']:
                try:
                    created_dt = datetime.strptime(created, fmt)
                    age_days = (datetime.now() - created_dt).days
                    if age_days < 180:
                        st.markdown(f'''
                        <div dir="rtl" style="
                            background:#1E3A8A; color:#F1F5F9; padding:12px;
                            border-radius:8px; border-right:4px solid #F59E0B;
                            margin-bottom:12px;
                            font-family:'Noto Sans Arabic','Tajawal',sans-serif;
                        ">
                        ℹ️ <strong>تنبيه:</strong> حساب جديد نسبياً ({age_days} يوم) - بيانات الجنسية قد تكون أقل موثوقية
                        </div>
                        ''', unsafe_allow_html=True)
                    break
                except ValueError:
                    continue
        except Exception:
            pass

    col_a, col_b = st.columns([1, 2])

    with col_a:
        country = result.get('country', 'غير متوفرة')
        flag = result.get('country_flag', '')
        country_ar = COUNTRY_AR.get(country, country)

        residence = result.get('country_original')
        show_residence = result.get('is_expatriate', False)

        # 🎟️ بطاقة الجنسية
        st.markdown(f"""
        <div class="country-card" dir="rtl">
            <div style="color: #FCD34D; font-size: 0.85rem; font-weight: 700; margin-bottom: 0.5rem;">🎟️ الجنسية</div>
            <div class="country-flag">{flag}</div>
            <div class="country-name">{country_ar}</div>
            <div style="color: #93C5FD; font-size: 0.9rem; margin-top: 0.5rem;">{country}</div>
        </div>
        """, unsafe_allow_html=True)

        # 📍 بطاقة الإقامة
        if show_residence and residence:
            residence_flag = ''
            for emoji, c in FLAG_EMOJI_TO_COUNTRY.items():
                if c == residence:
                    residence_flag = emoji
                    break
            residence_ar = COUNTRY_AR.get(residence, residence)
            st.markdown(f"""
            <div class="residence-card" dir="rtl">
                <div style="color: #D1FAE5; font-size: 0.85rem; font-weight: 700; margin-bottom: 0.5rem;">📍 موقع الإقامة</div>
                <div style="font-size: 3rem; line-height: 1; margin-bottom: 0.5rem;">{residence_flag}</div>
                <div style="color: #F1F5F9; font-size: 1.4rem; font-weight: 700;">{residence_ar}</div>
                <div style="color: #A7F3D0; font-size: 0.85rem; margin-top: 0.5rem;">{residence}</div>
            </div>
            """, unsafe_allow_html=True)
            expat_conf = result.get('expat_confidence', 0)
            st.markdown(f'<div style="text-align: center;" dir="rtl"><span class="expat-badge">🛂 مغترب ({expat_conf}%)</span></div>', unsafe_allow_html=True)

        # 🏙️ بطاقة المنطقة الذكية
        region_info = result.get('region_info')
        if region_info:
            is_estimate = region_info.get('is_estimate', False)
            estimate_note = region_info.get('estimate_note', '')
            card_class = 'region-card-estimate' if is_estimate else 'region-card'
            title = '🏙️ المنطقة (تقديري)' if is_estimate else '🏙️ المنطقة / المدينة'
            badge = f'<div class="estimate-badge">🟡 {estimate_note}</div>' if is_estimate else ''
            st.markdown(f"""
            <div class="{card_class}" dir="rtl">
                <div style="color: #FED7AA; font-size: 0.85rem; font-weight: 700; margin-bottom: 0.5rem;">{title}</div>
                <div style="font-size: 3rem; line-height: 1; margin-bottom: 0.5rem;">📍</div>
                <div style="color: #F1F5F9; font-size: 1.4rem; font-weight: 700;">{region_info.get('region_ar')}</div>
                <div style="color: #FFEDD5; font-size: 0.85rem; margin-top: 0.5rem;">{region_info.get('region_en')} · {region_info.get('confidence', 90)}%</div>
                {badge}
            </div>
            """, unsafe_allow_html=True)

        # شارة الثقة
        conf = result.get('confidence', 0)
        if conf >= 90:
            conf_class, conf_text = "confidence-high", "موثوق جداً"
        elif conf >= 70:
            conf_class, conf_text = "confidence-medium", "موثوق"
        else:
            conf_class, conf_text = "confidence-low", "تحقق يدوي"
        st.markdown(f'<div style="text-align: center; margin: 1rem 0;" dir="rtl"><span class="{conf_class}">🛡️ {conf_text} ({conf}%)</span></div>', unsafe_allow_html=True)

    with col_b:
        nickname = result.get('nickname', result.get('username'))
        st.markdown(f"""
        <div class="result-card" dir="rtl">
            <h2 style="color: #F59E0B; margin: 0 0 0.5rem 0; font-weight: 900;">{nickname}</h2>
            <p style="color: #94A3B8; font-size: 1.1rem; margin: 0;">@{result.get('username')}</p>
            <p style="color: #CBD5E1; margin-top: 1.2rem; line-height: 1.8; font-size: 1rem;">{result.get('bio') or '— لا يوجد وصف —'}</p>
            <div style="margin-top: 1rem; padding-top: 1rem; border-top: 1px solid rgba(245,158,11,0.2);">
                <span style="color: #94A3B8;">🗣️ اللغة: <strong style="color: #F59E0B;">{_format_language_label(result.get('language'))}</strong></span>
                <span style="color: #94A3B8; margin-right: 1.5rem;">📅 الإنشاء: <strong style="color: #F59E0B;">{result.get('created') or 'غير متوفر'}</strong></span>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # 📊 الإحصائيات
    st.markdown('<h3 dir="rtl" style="text-align:right; color: #F59E0B; margin-top: 1.5rem;">📊 الإحصائيات</h3>', unsafe_allow_html=True)
    stat_cols = st.columns(5)
    stats = [
        ("👥", "المتابعون", result.get('followers', 0)),
        ("❤️", "الإعجابات", result.get('hearts', 0)),
        ("🎬", "الفيديوهات", result.get('videos', 0)),
        ("➕", "يتابع", result.get('following', 0)),
        ("👫", "الأصدقاء", result.get('friends', 0)),
    ]
    for i, (icon, label, value) in enumerate(stats):
        with stat_cols[i]:
            if value >= 1_000_000:
                formatted = f"{value/1_000_000:.1f}M"
            elif value >= 1_000:
                formatted = f"{value/1_000:.1f}K"
            else:
                formatted = f"{value:,}"
            st.markdown(f'<div class="stat-card" dir="rtl"><div style="font-size: 2rem;">{icon}</div><div class="stat-number">{formatted}</div><div class="stat-label">{label}</div></div>', unsafe_allow_html=True)

    # 🧭 ✅ v2.1.7-Light-Fix3.1 - بطاقة النشاط الجغرافي + التحفّظ الشفّاف
    actual_residence = result.get('actual_residence')
    residence_confidence = result.get('residence_confidence', 0)
    residence_type = result.get('residence_type', '—')
    previous_residence = result.get('previous_residence')
    videos_analyzed = result.get('videos_analyzed', 0)
    region_distribution = result.get('region_distribution') or {}

    if actual_residence or residence_confidence > 0:
        # ترجمة الدولة
        actual_ar = COUNTRY_AR.get(actual_residence, actual_residence) if actual_residence else '—'
        # علم
        actual_flag = ''
        for emoji, c in FLAG_EMOJI_TO_COUNTRY.items():
            if c == actual_residence:
                actual_flag = emoji
                break

        # لون الثقة
        if residence_confidence >= 85:
            conf_color = '#10B981'  # أخضر
            conf_label = 'موثوقة جداً'
        elif residence_confidence >= 70:
            conf_color = '#F59E0B'  # برتقالي
            conf_label = 'موثوقة'
        else:
            conf_color = '#94A3B8'  # رمادي
            conf_label = 'منخفضة'

        # نوع الإقامة
        type_map = {
            'ثابت': '🟢 ثابت',
            'مستقر حديثاً': '🟡 مستقر حديثاً',
            'مستقر': '🟢 مستقر',
            'مسافر/متنقّل': '🟠 مسافر/متنقّل',
        }
        type_display = type_map.get(residence_type, f'⚪ {residence_type}')

        # توزيع المناطق
        dist_html = ''
        if region_distribution:
            items = []
            total = sum(region_distribution.values())
            for ctry, cnt in sorted(region_distribution.items(), key=lambda x: -x[1]):
                pct = int(round(cnt * 100 / total)) if total else 0
                ctry_ar = COUNTRY_AR.get(ctry, ctry)
                items.append(f'<span style="background:#1E3A8A;color:#F1F5F9;padding:4px 10px;border-radius:8px;margin-left:6px;display:inline-block;margin-bottom:4px;">{ctry_ar} ({ctry}): {cnt} ({pct}%)</span>')
            dist_html = ''.join(items)

        # سابقة الإقامة
        prev_html = ''
        if previous_residence:
            prev_ar = COUNTRY_AR.get(previous_residence, previous_residence)
            prev_html = f'<div style="margin-top:8px;color:#CBD5E1;"><strong style="color:#F59E0B;">🕒 إقامة سابقة:</strong> {prev_ar} ({previous_residence})</div>'

        st.markdown(f"""
        <div dir="rtl" style="
            background:#0F172A; padding:18px; border-radius:12px;
            border-right:4px solid {conf_color}; margin-top:18px;
            font-family:'Noto Sans Arabic','Tajawal',sans-serif;
            color:#F1F5F9;
        ">
            <h3 style="color:#F59E0B; margin:0 0 12px 0; font-weight:900;">🧭 النشاط الجغرافي (Fix3)</h3>
            <div style="display:flex; flex-wrap:wrap; gap:24px; align-items:center;">
                <div style="font-size:3.5rem; line-height:1;">{actual_flag}</div>
                <div style="flex:1; min-width:200px;">
                    <div style="color:#94A3B8; font-size:0.85rem; margin-bottom:4px;">📍 الإقامة الفعلية</div>
                    <div style="font-size:1.6rem; font-weight:800; color:#F1F5F9;">{actual_ar}</div>
                    <div style="color:#93C5FD; font-size:0.95rem; margin-top:4px;">{actual_residence or '—'}</div>
                </div>
                <div style="text-align:center;">
                    <div style="color:#94A3B8; font-size:0.85rem;">درجة الثقة</div>
                    <div style="font-size:2rem; font-weight:900; color:{conf_color};">{residence_confidence}%</div>
                    <div style="font-size:0.85rem; color:{conf_color};">{conf_label}</div>
                </div>
                <div style="text-align:center;">
                    <div style="color:#94A3B8; font-size:0.85rem;">النوع</div>
                    <div style="font-size:1.1rem; font-weight:700; margin-top:4px;">{type_display}</div>
                </div>
            </div>
            <div style="margin-top:14px; padding-top:12px; border-top:1px solid rgba(245,158,11,0.2);">
                <div style="color:#CBD5E1;"><strong style="color:#F59E0B;">📊 الفيديوهات المُحلَّلة:</strong> {videos_analyzed}</div>
                <div style="margin-top:8px; color:#CBD5E1;"><strong style="color:#F59E0B;">🗺️ توزيع المناطق:</strong></div>
                <div style="margin-top:6px;">{dist_html or '<span style=\"color:#94A3B8;\">—</span>'}</div>
                {prev_html}
            </div>
        </div>
        """, unsafe_allow_html=True)

        # 🛡️ بطاقة التحفّظ الشفّاف - ✅ Fix3.1
        st.markdown(f"""
        <div dir="rtl" style="
            background:#F1F5F9; padding:16px; border-radius:10px;
            border-right:4px solid #F59E0B; margin-top:12px; color:#0F172A;
            font-family:'Noto Sans Arabic','Tajawal',sans-serif;
        ">
            <h4 style="margin:0 0 8px 0; color:#0F172A;">🛡️ تحفّظ شفّاف على دقّة الإقامة</h4>
            <p style="margin:0; line-height:1.8; font-size:0.95rem;">
                موقع الإقامة المُستنتَج يعتمد على آخر <strong>{videos_analyzed} مقاطع منشورة</strong> + التسلسل الزمني عبر TikWM.
                إذا توقّف صاحب الحساب عن النشر، أو نشر من بلد سفر مؤقّت، أو حذف مقاطع، قد لا يعكس الموقع الحقيقي للإقامة الراهنة.
                نلتزم بعدم استخدام IP أو منصّات خارجية حفاظاً على الخصوصية وميثاق المشروع.
            </p>
        </div>
        """, unsafe_allow_html=True)

    # 🗺️ ✅ v2.1.7-Light-Fix3.2-Map - خريطة الدول التفاعلية
    if region_distribution:
        try:
            from streamlit_folium import st_folium
            _map = render_country_choropleth(region_distribution, actual_residence)
            if _map is not None:
                st.markdown("""
                <div dir="rtl" style="
                    background:#0F172A; padding:14px 18px; border-radius:12px 12px 0 0;
                    border-right:4px solid #F59E0B; margin-top:18px; color:#F1F5F9;
                    font-family:'Noto Sans Arabic','Tajawal',sans-serif;
                ">
                    <h3 style="color:#F59E0B; margin:0 0 6px 0; font-weight:900;">🗺️ خريطة مناطق الفيديوهات</h3>
                    <p style="margin:0; color:#CBD5E1; font-size:0.9rem;">
                        تظليل الدول المُستخرجة من آخر {n} مقاطع — مستوى الدولة فقط، لا مدن ولا إحداثيات دقيقة.
                    </p>
                </div>
                """.format(n=videos_analyzed or len(region_distribution)), unsafe_allow_html=True)

                st_folium(_map, height=400, width=None, returned_objects=[])

                # 🛡️ تحفّظ خاص بالخريطة - Fix3.2
                st.markdown("""
                <div dir="rtl" style="
                    background:#F1F5F9; padding:14px 16px; border-radius:0 0 10px 10px;
                    border-right:4px solid #F59E0B; margin-top:-4px; color:#0F172A;
                    font-family:'Noto Sans Arabic','Tajawal',sans-serif;
                ">
                    <strong>🛡️ تنبيه:</strong> الخريطة تعرض <strong>دول الفيديوهات</strong>، وليست بالضرورة <strong>دولة الإقامة الحالية</strong>.
                    لا تستخدم IP ولا Geocoding خارجي — مصدر البيانات: TikWM حصراً ضمن الميثاق.
                </div>
                """, unsafe_allow_html=True)
        except Exception as _map_err:
            st.markdown(f"""
            <div dir="rtl" style="
                background:#FEF3C7; padding:12px; border-radius:8px; color:#0F172A;
                border-right:4px solid #F59E0B;
                font-family:'Noto Sans Arabic','Tajawal',sans-serif;
            ">
                ℹ️ الخريطة غير متاحة حالياً (تحقّق من حزمة folium في requirements.txt).
            </div>
            """, unsafe_allow_html=True)

    # 🔧 تفاصيل تقنية للمطورين - ✅ v2.1.5 مغلقة افتراضياً
    with st.expander("🔧 تفاصيل تقنية للمطورين", expanded=False):
        user_id = result.get('user_id', '—')
        sec_uid = result.get('sec_uid', '—')
        if sec_uid != '—' and len(sec_uid) > 20:
            sec_uid_display = f"{sec_uid[:15]}...{sec_uid[-8:]}"
        else:
            sec_uid_display = sec_uid

        st.markdown(f"""
        <div class="tech-details-card" dir="rtl">
            <div class="tech-grid">
                <div class="tech-item">
                    <div class="tech-label">🆔 معرّف المستخدم</div>
                    <div class="tech-value">{user_id}</div>
                </div>
                <div class="tech-item">
                    <div class="tech-label">🔐 SecUID</div>
                    <div class="tech-value">{sec_uid_display}</div>
                </div>
                <div class="tech-item">
                    <div class="tech-label">📅 تاريخ الإنشاء</div>
                    <div class="tech-value">{result.get('created', '—')}</div>
                </div>
                <div class="tech-item">
                    <div class="tech-label">⚡ زمن الاستجابة</div>
                    <div class="tech-value">{result.get('fetch_time', 0)} ثانية</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)


def process_bulk(usernames, progress_bar, status_text):
    """معالجة قائمة حسابات بدفعات - ✅ v2.1.1 مع معالجة الأخطاء"""
    results = []
    total = len(usernames)
    errors_log = []
    for i, u in enumerate(usernames, 1):
        status_text.markdown(
            f'<div dir="rtl" style="font-family:Noto Sans Arabic,Tajawal,sans-serif;">'
            f'🔍 جاري معالجة {i}/{total}: <code>@{u}</code></div>',
            unsafe_allow_html=True
        )
        # ✅ تطوير #3 v2.1.1 - try/except لمنع توقف الدفعة
        try:
            result = lookup_user(u)
            results.append(result)
        except Exception as e:
            error_result = {
                'success': False,
                'username': u,
                'error': f'خطأ معالجة: {str(e)[:100]}',
                'reason': 'processing_exception',
            }
            results.append(error_result)
            errors_log.append({'username': u, 'error': str(e)[:200]})
        progress_bar.progress(i / total)
        if i < total:
            # ✅ v2.1.7-Light-Fix2 - inter-account: 2.5-4.0ث (أو 4.0-6.4ث في burst)
            _inter = random.uniform(2.5, 4.0)
            if _is_in_burst_mode():
                _inter *= 1.6
            time.sleep(_inter)
    if errors_log:
        st.session_state['_bulk_errors'] = errors_log
    return results


def results_to_dataframe(results):
    """تحويل النتائج لـ DataFrame جميل - ✅ v2.1.3 إضافة عمود رابط الحساب"""
    try:
        import pandas as pd
    except ImportError:
        return None
    rows = []
    for r in results:
        username = r.get('username', '?')
        tiktok_url = f"https://www.tiktok.com/@{username}"
        
        if not r.get('success'):
            rows.append({
                'الحساب': f"@{username}",
                '🔗 رابط الحساب': tiktok_url,
                'الحالة': '❌ فشل',
                'السبب': r.get('error', '—'),
            })
            continue
        country = r.get('country', '—')
        residence = r.get('country_original') if r.get('is_expatriate') else None
        region = r.get('region_info', {}).get('region_ar') if r.get('region_info') else None
        rows.append({
            'الحساب': f"@{username}",
            '🔗 رابط الحساب': tiktok_url,  # ✅ عمود جديد v2.1.3
            'الاسم': r.get('nickname', '—'),
            'الجنسية': COUNTRY_AR.get(country, country),
            'العلم': r.get('country_flag', ''),
            'الإقامة': COUNTRY_AR.get(residence, residence) if residence else '—',
            'المنطقة': region or '—',
            'المتابعون': r.get('followers', 0),
            'الإعجابات': r.get('hearts', 0),
            'الفيديوهات': r.get('videos', 0),
            'يتابع': r.get('following', 0),
            'البايو': r.get('bio', '—'),
            'اللغة': r.get('language', '—'),
            'الإنشاء': r.get('created', '—'),
            'الثقة %': r.get('confidence', 0),
            'User ID': r.get('user_id', '—'),
        })
    return pd.DataFrame(rows)


def export_to_excel(df):
    """تصدير DataFrame إلى Excel bytes"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "نتائج بَصِير"
    ws.sheet_view.rightToLeft = True

    # رأس الجدول
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="F59E0B")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # البيانات
    for row_num, row in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # عرض الأعمدة
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 18

    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════════
# 🎨 الواجهة الرئيسية بالـ Tabs
# ═══════════════════════════════════════════════════════════════
tab1, tab2, tab3 = st.tabs(["🔍 بحث فردي", "📋 بحث جماعي", "📊 استيراد Excel"])

# ─────── تاب 1: البحث الفردي ───────
with tab1:
    col1, col2, col3 = st.columns([1, 3, 1])
    with col2:
        username_input = st.text_input(
            "اسم المستخدم على TikTok",
            placeholder="مثال: aboflah أو رابط TikTok",
            help="أدخل اسم المستخدم بدون @ أو الصق رابط TikTok",
            key="single_search"
        )
        # ✅ v2.1.5 - تعطيل زر البحث أثناء المعالجة
        _is_busy_single = st.session_state.get('_single_processing', False)
        search_btn = st.button(
            "⏳ جاري المعالجة..." if _is_busy_single else "🔍 ابحث الآن",
            type="primary", use_container_width=True, key="btn_single",
            disabled=_is_busy_single,
        )

    # ✅ v2.1.5 - تحقق صريح قبل البحث الفردي + تعطيل الزر
    if search_btn:
        if not username_input or not username_input.strip():
            st.markdown(
                '<div dir="rtl" style="background: rgba(59,130,246,0.12); padding: 0.9rem 1rem; border-radius: 8px; border-right: 4px solid #3B82F6; color: #BFDBFE; font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;">⚠️ يُرجى إدخال اسم مستخدم أو رابط TikTok قبل الضغط على زر البحث.</div>',
                unsafe_allow_html=True
            )
        else:
            st.session_state['_single_processing'] = True
            try:
                with st.spinner("🔍 جاري البحث..."):
                    result = lookup_user(username_input)
                display_single_result(result)
            finally:
                st.session_state['_single_processing'] = False

# ─────── تاب 2: البحث الجماعي ───────
with tab2:
    st.markdown('<div dir="rtl" style="font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif; color: #94A3B8; margin-bottom: 1rem;">📋 الصق قائمة الحسابات (واحد في كل سطر). يدعم: @user أو رابط TikTok أو اسم فقط.</div>', unsafe_allow_html=True)

    bulk_input = st.text_area(
        "قائمة الحسابات",
        placeholder="aboflah\nmrbeast\ncristiano\nhttps://tiktok.com/@drake",
        height=200,
        max_chars=5000,
    )

    # ══════════════════════════════════════════════════════════
    # ✅ v2.1.7-Light-Fix2 — Auto-Limit Smart (Safe Hard Limit: 50)
    # ══════════════════════════════════════════════════════════
    SAFE_LIMIT = 50  # الحد الأقصى الآمن (توصية معتمدة BSR-V217L-SAFE-LIMIT)
    raw_lines = [ln.strip() for ln in (bulk_input or '').strip().split('\n') if ln.strip()]
    detected_count = len(raw_lines)
    max_accounts = min(detected_count, SAFE_LIMIT) if detected_count > 0 else 0

    # تحديد الشارة والوقت التقريبي (مع throttling الجديد Fix2: ~4.2ث/حساب)
    if detected_count == 0:
        badge_html = (
            '<div dir="rtl" style="font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;'
            ' background:#0F172A; padding:12px 16px; border-radius:10px;'
            ' border-right:4px solid #94A3B8; color:#94A3B8;">'
            '— أدخل قائمة الحسابات أعلاه لبدء التحليل</div>'
        )
    elif detected_count <= 10:
        eta_sec = round(detected_count * 4.2)
        badge_html = (
            f'<div dir="rtl" style="font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;'
            f' background:#0F172A; padding:12px 16px; border-radius:10px;'
            f' border-right:4px solid #15803D; color:#F1F5F9;">'
            f'<span style="background:#15803D; color:white; padding:3px 12px;'
            f' border-radius:14px; font-weight:bold;">🟢 سريع</span>'
            f' &nbsp; {detected_count} حساب · وقت متوقع ~{eta_sec} ثانية</div>'
        )
    elif detected_count <= 30:
        eta_min = round(detected_count * 4.2 / 60, 1)
        badge_html = (
            f'<div dir="rtl" style="font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;'
            f' background:#0F172A; padding:12px 16px; border-radius:10px;'
            f' border-right:4px solid #F59E0B; color:#F1F5F9;">'
            f'<span style="background:#F59E0B; color:#0F172A; padding:3px 12px;'
            f' border-radius:14px; font-weight:bold;">🟡 متوسط</span>'
            f' &nbsp; {detected_count} حساب · وقت متوقع ~{eta_min} دقيقة</div>'
        )
    elif detected_count <= SAFE_LIMIT:
        eta_min = round(detected_count * 4.2 / 60, 1)
        badge_html = (
            f'<div dir="rtl" style="font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;'
            f' background:#0F172A; padding:12px 16px; border-radius:10px;'
            f' border-right:4px solid #EA580C; color:#F1F5F9;">'
            f'<span style="background:#EA580C; color:white; padding:3px 12px;'
            f' border-radius:14px; font-weight:bold;">🟠 مكثّف</span>'
            f' &nbsp; {detected_count} حساب · وقت متوقع ~{eta_min} دقيقة</div>'
        )
    else:
        eta_min = round(SAFE_LIMIT * 4.2 / 60, 1)
        cut = detected_count - SAFE_LIMIT
        badge_html = (
            f'<div dir="rtl" style="font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;'
            f' background:#0F172A; padding:12px 16px; border-radius:10px;'
            f' border-right:4px solid #DC2626; color:#F1F5F9;">'
            f'<span style="background:#DC2626; color:white; padding:3px 12px;'
            f' border-radius:14px; font-weight:bold;">🔴 تجاوز الحد الآمن</span>'
            f' &nbsp; {detected_count} حساب — سيُعالَج أول <strong>{SAFE_LIMIT}</strong> فقط (~{eta_min} دقيقة)'
            f'<br><span style="color:#FCA5A5; font-size:0.85em;">'
            f'⚠️ سيُترك {cut} حساب جانباً حمايةً من rate-limit — رشّح الدفعة التالية بالباقي</span></div>'
        )

    st.markdown(badge_html, unsafe_allow_html=True)

    # زر المعالجة بعرض كامل
    _is_busy_bulk = st.session_state.get('_bulk_processing', False)
    btn_label = (
        "⏳ جاري المعالجة..." if _is_busy_bulk
        else (f"🚀 معالجة {max_accounts} حساب الآن" if max_accounts > 0
              else "🚀 معالجة الجميع")
    )
    bulk_btn = st.button(
        btn_label,
        type="primary", use_container_width=True, key="btn_bulk",
        disabled=_is_busy_bulk or max_accounts == 0,
    )

    if bulk_btn and bulk_input.strip():
        st.session_state['_bulk_processing'] = True
        usernames = [line.strip() for line in bulk_input.strip().split('\n') if line.strip()]
        usernames = usernames[:max_accounts]
        # تنظيف
        cleaned = []
        for u in usernames:
            u = u.strip().lstrip('@')
            u = re.sub(r'https?://(?:www\.)?tiktok\.com/@?', '', u)
            u = u.split('?')[0].split('/')[0]
            if u:
                cleaned.append(u)

        st.markdown(f'<div dir="rtl" style="color: #94A3B8;">🎯 سيتم معالجة <strong style="color: #F59E0B;">{len(cleaned)}</strong> حساب</div>', unsafe_allow_html=True)

        progress_bar = st.progress(0.0)
        status_text = st.empty()

        results = process_bulk(cleaned, progress_bar, status_text)
        status_text.markdown('<div dir="rtl" style="color: #10B981; font-weight: 700; font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;">✅ اكتملت المعالجة!</div>', unsafe_allow_html=True)
        st.session_state['bulk_results'] = results

        st.session_state['_bulk_processing'] = False
        # ✅ v2.1.5 - صندوق ملخص النجاح/الفشل
        total_count = len(results)
        success_count = sum(1 for r in results if r.get('success'))
        fail_count = total_count - success_count
        success_pct = round((success_count / total_count * 100), 1) if total_count else 0
        st.markdown(
            f'''
            <div dir="rtl" style="background:#0F172A; color:#F1F5F9; padding:18px 20px; border-radius:12px; border-right:5px solid #F59E0B; margin:18px 0; font-family:'Noto Sans Arabic','Tajawal',sans-serif;">
                <h4 style="color:#F59E0B; margin:0 0 12px 0;">📊 ملخص المعالجة الجماعية</h4>
                <div style="display:flex; gap:24px; flex-wrap:wrap;">
                    <div>✅ <strong>ناجح:</strong> <span style="color:#10B981;">{success_count}</span></div>
                    <div>❌ <strong>فاشل:</strong> <span style="color:#EF4444;">{fail_count}</span></div>
                    <div>📌 <strong>الإجمالي:</strong> <span style="color:#F59E0B;">{total_count}</span></div>
                    <div>📈 <strong>نسبة النجاح:</strong> <span style="color:#3B82F6;">{success_pct}%</span></div>
                </div>
            </div>
            ''',
            unsafe_allow_html=True
        )

        # ✅ v2.1.5 - جدول أخطاء منفصل للفاشلين فقط
        if fail_count > 0:
            with st.expander(f"⚠️ عرض الحسابات الفاشلة ({fail_count})", expanded=False):
                try:
                    import pandas as _pd
                    fail_rows = [
                        {
                            'الحساب': f"@{r.get('username','?')}",
                            'سبب الفشل': r.get('error', '—'),
                        }
                        for r in results if not r.get('success')
                    ]
                    fail_df = _pd.DataFrame(fail_rows)
                    st.dataframe(fail_df, use_container_width=True, hide_index=True)
                except Exception:
                    for r in results:
                        if not r.get('success'):
                            st.markdown(
                                f'<div dir="rtl" style="color:#FCA5A5; font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;">❌ @{r.get("username","?")} — {r.get("error","—")}</div>',
                                unsafe_allow_html=True
                            )

        # عرض الجدول - ✅ v2.1.3 رابط قابل للنقر
        df = results_to_dataframe(results)
        if df is not None:
            st.markdown('<h3 dir="rtl" style="color: #F59E0B; margin-top: 1.5rem;">📊 جدول النتائج</h3>', unsafe_allow_html=True)
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "🔗 رابط الحساب": st.column_config.LinkColumn(
                        "🔗 رابط الحساب",
                        help="اضغط لفتح الحساب في TikTok",
                        display_text="🔗 فتح",
                        width="small",
                    ),
                },
            )

            # تصدير Excel
            excel_bytes = export_to_excel(df)
            if excel_bytes:
                st.download_button(
                    label="📥 تنزيل النتائج (Excel)",
                    data=excel_bytes,
                    file_name=f"baseer_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

# ─────── تاب 3: استيراد Excel - ✅ v2.1.4 ───────
with tab3:
    st.markdown('<div dir="rtl" style="color: #94A3B8; margin-bottom: 1rem;">📊 ارفع ملف Excel يحتوي على عمود <code>username</code> أو <code>url</code></div>', unsafe_allow_html=True)

    # ✅ v2.1.4 - زر تحميل نموذج Excel جاهز
    st.markdown('<div dir="rtl" style="background:#1E3A8A; color:#F1F5F9; padding:14px; border-radius:10px; border-right:4px solid #F59E0B; margin-bottom:14px; font-family:Noto Sans Arabic,Tajawal,sans-serif;">💡 <strong>ليس لديك ملف جاهز؟</strong> حمّل النموذج أدناه، املأه بالحسابات، ثم ارفعه هنا.</div>', unsafe_allow_html=True)
    
    col_template1, col_template2, col_template3 = st.columns([1, 1, 1])
    
    with col_template1:
        # ✅ إنشاء نموذج Excel جاهز للتحميل
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            template_wb = Workbook()
            template_ws = template_wb.active
            template_ws.title = "قالب بَصِير"
            template_ws.sheet_view.rightToLeft = True
            
            # رأس الجدول
            headers = ['username', 'الملاحظات (اختياري)']
            thin_border = Border(
                left=Side(style='thin', color='F59E0B'),
                right=Side(style='thin', color='F59E0B'),
                top=Side(style='thin', color='F59E0B'),
                bottom=Side(style='thin', color='F59E0B'),
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = template_ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12, name='Tajawal')
                cell.fill = PatternFill("solid", fgColor="F59E0B")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # أمثلة توضيحية
            examples = [
                ['abu_ali838', 'مثال حساب مباشر'],
                ['@rsn0077', 'مثال برمز @ - سيلّخص تلقائياً'],
                ['https://www.tiktok.com/@aboflah', 'مثال رابط كامل'],
                ['khaby.lame', ''],
                ['samira.tn', ''],
                ['', ''],  # صفوف فارغة للتعبئة
                ['', ''],
                ['', ''],
                ['', ''],
                ['', ''],
            ]
            
            for row_num, row_data in enumerate(examples, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = template_ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = Font(name='Tajawal', size=11)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = thin_border
            
            # عرض الأعمدة
            template_ws.column_dimensions['A'].width = 35
            template_ws.column_dimensions['B'].width = 45
            
            # ورقة تعليمات
            instructions_ws = template_wb.create_sheet("التعليمات")
            instructions_ws.sheet_view.rightToLeft = True
            instructions = [
                ['📖 دليل استخدام النموذج'],
                [''],
                ['1. أدخل حسابات TikTok في عمود "username"'],
                ['2. الصيغ المدعومة:'],
                ['   - abu_ali838 (بدون @)'],
                ['   - @abu_ali838 (مع @)'],
                ['   - https://www.tiktok.com/@abu_ali838 (رابط كامل)'],
                ['3. الحد الأقصى: 50 حساب لكل ملف'],
                ['4. احفظ الملف بصيغة .xlsx'],
                ['5. ارفعه في التبويب أعلاه'],
                [''],
                ['⚠️ ملاحظات مهمة:'],
                ['- احذف صفوف الأمثلة قبل الرفع (اختياري)'],
                ['- لا تغير اسم عمود "username"'],
                ['- الصفوف الفارغة ستُتجاهل تلقائياً'],
                [''],
                ['🔖 الإصدار: Baseer v2.1.4'],
                ['📅 التاريخ: 2026-06-05'],
            ]
            for row_num, row_data in enumerate(instructions, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = instructions_ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = Font(name='Tajawal', size=12, bold=(row_num == 1))
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            instructions_ws.column_dimensions['A'].width = 65
            
            template_buffer = io.BytesIO()
            template_wb.save(template_buffer)
            template_buffer.seek(0)
            
            st.download_button(
                label="📄 تحميل النموذج",
                data=template_buffer.getvalue(),
                file_name="baseer_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_template",
            )
        except Exception as e:
            st.error(f"تعذّر إنشاء النموذج: {str(e)[:100]}")
    
    with col_template2:
        # ✅ نموذج CSV بديل
        try:
            csv_template = "username,الملاحظات\nabu_ali838,مثال\n@rsn0077,مثال برمز\nhttps://www.tiktok.com/@aboflah,رابط كامل\n,\n,\n,\n,\n,\n,\n"
            csv_bytes = csv_template.encode('utf-8-sig')  # BOM لدعم العربية في Excel
            st.download_button(
                label="📝 تحميل CSV",
                data=csv_bytes,
                file_name="baseer_template.csv",
                mime="text/csv",
                use_container_width=True,
                key="download_csv",
            )
        except Exception:
            pass
    
    with col_template3:
        st.markdown('<div dir="rtl" style="color:#94A3B8; padding-top:8px; font-family:Noto Sans Arabic,sans-serif;">✨ يمكنك رفع .xlsx أو .xls أو .csv</div>', unsafe_allow_html=True)

    st.divider()
    
    uploaded_file = st.file_uploader(
        "اختر ملف (Excel أو CSV)",
        type=['xlsx', 'xls', 'csv'],
        key="excel_upload",
        help="الصيغ المدعومة: xlsx, xls, csv — الحد الأقصى 50 حساب",
    )

    if uploaded_file is not None:
        try:
            import pandas as pd
            
            # ✅ v2.1.4 - دعم CSV + Excel
            file_ext = uploaded_file.name.lower().split('.')[-1]
            if file_ext == 'csv':
                df_input = pd.read_csv(uploaded_file, encoding='utf-8-sig')
            else:
                df_input = pd.read_excel(uploaded_file)
            
            # ✅ v2.1.4 - تنظيف أسماء الأعمدة
            df_input.columns = [str(c).strip() for c in df_input.columns]
            
            st.markdown(f'<div dir="rtl" style="color: #10B981; font-family:Noto Sans Arabic,sans-serif;">✅ تم رفع الملف بنجاح ({len(df_input)} سطر ، {len(df_input.columns)} عمود)</div>', unsafe_allow_html=True)
            
            # عرض أول 10 صفوف
            st.markdown('<div dir="rtl" style="color:#F59E0B; margin-top:12px;">👀 معاينة الملف:</div>', unsafe_allow_html=True)
            st.dataframe(df_input.head(10), use_container_width=True)

            # ✅ v2.1.4 - بحث جوهري في أسماء الأعمدة (غير حساس للحالة)
            col_name = None
            preferred_cols = ['username', 'user', 'url', 'link', 'account', 'tiktok', 'handle']
            
            # بحث دقيق أولاً
            for col in df_input.columns:
                col_lower = str(col).lower().strip()
                if col_lower in preferred_cols:
                    col_name = col
                    break
            
            # بحث جزئي إذا لم يوجد
            if not col_name:
                for col in df_input.columns:
                    col_lower = str(col).lower().strip()
                    if any(kw in col_lower for kw in preferred_cols):
                        col_name = col
                        break
            
            # ✅ v2.1.4 - اختيار يدوي إذا لم يتم الاكتشاف التلقائي
            if not col_name and len(df_input.columns) > 0:
                st.warning('⚠️ لم أتمكّن من العثور على عمود الحسابات تلقائياً - اختره يدوياً:')
                col_name = st.selectbox(
                    "اختر العمود الذي يحوي الحسابات:",
                    options=list(df_input.columns),
                    key="manual_col_select",
                )

            if col_name:
                # ✅ v2.1.4 - تصفية الصفوف الفارغة أولاً
                valid_rows = df_input[col_name].dropna()
                valid_rows = valid_rows[valid_rows.astype(str).str.strip() != '']
                
                st.markdown(
                    f'<div dir="rtl" style="background:#0F172A; color:#F1F5F9; padding:12px; border-radius:8px; border-right:4px solid #10B981; font-family:Noto Sans Arabic,sans-serif;">'
                    f'🎯 <strong>العمود المُكتشَف:</strong> <span style="color:#F59E0B;">{col_name}</span><br>'
                    f'📊 <strong>عدد الصفوف الصالحة:</strong> {len(valid_rows)} حساب'
                    f'</div>',
                    unsafe_allow_html=True
                )
                
                if len(valid_rows) == 0:
                    st.error("❌ لا توجد حسابات صالحة في العمود المحدد")
                else:
                    # ✅ v2.1.5 - تعطيل زر Excel أثناء المعالجة
                    _is_busy_excel = st.session_state.get('_excel_processing', False)
                    process_excel_btn = st.button(
                        "⏳ جاري المعالجة..." if _is_busy_excel else f"🚀 معالجة {min(len(valid_rows), 50)} حساب",
                        type="primary",
                        key="btn_excel",
                        use_container_width=True,
                        disabled=_is_busy_excel,
                    )

                if process_excel_btn:
                    st.session_state['_excel_processing'] = True
                    usernames = df_input[col_name].dropna().astype(str).tolist()[:50]
                    cleaned = []
                    for u in usernames:
                        u = u.strip().lstrip('@')
                        u = re.sub(r'https?://(?:www\.)?tiktok\.com/@?', '', u)
                        u = u.split('?')[0].split('/')[0]
                        if u:
                            cleaned.append(u)

                    progress = st.progress(0.0)
                    status = st.empty()
                    results = process_bulk(cleaned, progress, status)
                    st.session_state['_excel_processing'] = False
                    status.markdown('<div dir="rtl" style="color: #10B981; font-weight: 700; font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;">✅ اكتملت المعالجة!</div>', unsafe_allow_html=True)

                    # ✅ v2.1.5 - صندوق ملخص + جدول أخطاء (استيراد Excel)
                    _total = len(results)
                    _ok = sum(1 for r in results if r.get('success'))
                    _fail = _total - _ok
                    _pct = round((_ok / _total * 100), 1) if _total else 0
                    st.markdown(
                        f'''
                        <div dir="rtl" style="background:#0F172A; color:#F1F5F9; padding:18px 20px; border-radius:12px; border-right:5px solid #F59E0B; margin:18px 0; font-family:'Noto Sans Arabic','Tajawal',sans-serif;">
                            <h4 style="color:#F59E0B; margin:0 0 12px 0;">📊 ملخص معالجة Excel</h4>
                            <div style="display:flex; gap:24px; flex-wrap:wrap;">
                                <div>✅ <strong>ناجح:</strong> <span style="color:#10B981;">{_ok}</span></div>
                                <div>❌ <strong>فاشل:</strong> <span style="color:#EF4444;">{_fail}</span></div>
                                <div>📌 <strong>الإجمالي:</strong> <span style="color:#F59E0B;">{_total}</span></div>
                                <div>📈 <strong>نسبة النجاح:</strong> <span style="color:#3B82F6;">{_pct}%</span></div>
                            </div>
                        </div>
                        ''',
                        unsafe_allow_html=True
                    )
                    if _fail > 0:
                        with st.expander(f"⚠️ عرض الحسابات الفاشلة ({_fail})", expanded=False):
                            try:
                                import pandas as _pd2
                                _fail_rows = [
                                    {'الحساب': f"@{r.get('username','?')}", 'سبب الفشل': r.get('error', '—')}
                                    for r in results if not r.get('success')
                                ]
                                st.dataframe(_pd2.DataFrame(_fail_rows), use_container_width=True, hide_index=True)
                            except Exception:
                                pass

                    df_results = results_to_dataframe(results)
                    if df_results is not None:
                        # ✅ v2.1.3 رابط قابل للنقر في Excel tab
                        st.dataframe(
                            df_results,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "🔗 رابط الحساب": st.column_config.LinkColumn(
                                    "🔗 رابط الحساب",
                                    help="اضغط لفتح الحساب في TikTok",
                                    display_text="🔗 فتح",
                                    width="small",
                                ),
                            },
                        )
                        excel_out = export_to_excel(df_results)
                        if excel_out:
                            st.download_button(
                                "📥 تنزيل النتائج (Excel)",
                                data=excel_out,
                                file_name=f"baseer_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
            else:
                st.markdown(
                    '<div dir="rtl" style="background: rgba(59,130,246,0.12); padding: 0.9rem 1rem; border-radius: 8px; border-right: 4px solid #3B82F6; color: #BFDBFE; font-family:\'Noto Sans Arabic\',\'Tajawal\',sans-serif;">⚠️ لم يتم العثور على عمود مناسب. التطبيق يقبل أي من: <code>username</code> | <code>url</code> | <code>name</code> | <code>account</code> | <code>الحساب</code> | <code>اسم المستخدم</code>.</div>',
                    unsafe_allow_html=True
                )

        except Exception as e:
            st.error(f"❌ خطأ في قراءة الملف: {e}")

    # قالب جاهز
    st.markdown('<div dir="rtl" style="margin-top: 2rem; color: #94A3B8;">💡 <strong>قالب جاهز:</strong> أنشئ ملف Excel بعمود واحد اسمه <code>username</code> وضع الحسابات سطراً واحداً.</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# Footer بسيط (لا تفاصيل تقنية)
# ═══════════════════════════════════════════════════════════════
st.markdown("""
<div style="text-align: center; margin-top: 3rem; padding: 1rem; color: #475569;" dir="rtl">
    <p style="font-size: 0.9rem;">🦅 <strong style="color: #F59E0B;">بَصِير</strong></p>
</div>
""", unsafe_allow_html=True)
